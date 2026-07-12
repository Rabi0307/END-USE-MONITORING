"""
build_excel.py
Builds the multi-sheet "Post-Disbursement End-Use of Funds (EUF) Monitoring"
workbook for LAP / Home Loan accounts, using end_use_engine.py to analyse
the borrower's bank statement.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
import pandas as pd
from datetime import datetime

from end_use_engine import run as run_engine

# --------------------------------------------------------------------------
# STYLE CONSTANTS
# --------------------------------------------------------------------------
FONT_NAME = "Arial"
NAVY = "1F3864"
LIGHT_BLUE = "DCE6F1"
YELLOW = "FFF2CC"
RED_FILL = "FFC7CE"
RED_FONT = "9C0006"
GREEN_FILL = "C6EFCE"
GREEN_FONT = "006100"
GREY = "F2F2F2"

TITLE_FONT = Font(name=FONT_NAME, size=14, bold=True, color="FFFFFF")
HEADER_FONT = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
SUBHEADER_FONT = Font(name=FONT_NAME, size=10, bold=True, color=NAVY)
NORMAL_FONT = Font(name=FONT_NAME, size=10)
INPUT_FONT = Font(name=FONT_NAME, size=10, color="0000FF")
NOTE_FONT = Font(name=FONT_NAME, size=9, italic=True, color="808080")

TITLE_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
INPUT_FILL = PatternFill("solid", fgColor=YELLOW)
SUBHEADER_FILL = PatternFill("solid", fgColor=LIGHT_BLUE)
STRIPE_FILL = PatternFill("solid", fgColor=GREY)

THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")


def style_title(ws, cell_range, text):
    ws.merge_cells(cell_range)
    c = ws[cell_range.split(":")[0]]
    c.value = text
    c.font = TITLE_FONT
    c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[c.row].height = 26


def style_header_row(ws, row, ncols, start_col=1):
    for i in range(start_col, start_col + ncols):
        c = ws.cell(row=row, column=i)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = 30


def autofit_columns(ws, df_or_none=None, min_width=10, max_width=55, extra=2):
    """Approximate Excel's 'AutoFit column width' using max content length per column."""
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            col = cell.column_letter
            length = len(str(cell.value))
            widths[col] = max(widths.get(col, 0), length)
    for col, length in widths.items():
        ws.column_dimensions[col].width = max(min_width, min(max_width, length + extra))


def write_df(ws, df, start_row, start_col=1, header_fill=True, currency_cols=None,
             pct_cols=None, date_cols=None, stripe=True):
    """Write a DataFrame as a formatted table starting at (start_row, start_col)."""
    currency_cols = currency_cols or []
    pct_cols = pct_cols or []
    date_cols = date_cols or []
    ncols = len(df.columns)
    for j, col in enumerate(df.columns):
        c = ws.cell(row=start_row, column=start_col + j, value=col)
    style_header_row(ws, start_row, ncols, start_col)

    for i, (_, r) in enumerate(df.iterrows()):
        row = start_row + 1 + i
        for j, col in enumerate(df.columns):
            cell = ws.cell(row=row, column=start_col + j, value=r[col])
            cell.font = NORMAL_FONT
            cell.border = BORDER
            cell.alignment = WRAP if isinstance(r[col], str) and len(str(r[col])) > 25 else LEFT
            if col in currency_cols:
                cell.number_format = '#,##0.00;(#,##0.00)'
            if col in pct_cols:
                cell.number_format = '0.0%'
            if col in date_cols:
                cell.number_format = 'dd-mmm-yyyy'
        if stripe and i % 2 == 1:
            for j in range(ncols):
                ws.cell(row=row, column=start_col + j).fill = STRIPE_FILL
    return start_row + 1 + len(df)  # next free row


def build_workbook(statement_path, loan_amount, loan_type, lender_hint,
                    borrower_name, loan_account_no, sanction_date,
                    purpose_declared, property_address, out_path,
                    window_days=90):

    df, disb_row, summary = run_engine(statement_path, loan_amount=loan_amount,
                                        loan_type=loan_type, lender_hint=lender_hint,
                                        window_days=window_days)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # =====================================================================
    # SHEET 1: INSTRUCTIONS
    # =====================================================================
    ws = wb.create_sheet("Instructions")
    style_title(ws, "A1:F1", "POST-DISBURSEMENT END-USE OF FUNDS (EUF) MONITORING — LAP / HOME LOAN")
    ws["A3"] = "Purpose"
    ws["A3"].font = SUBHEADER_FONT
    ws["A4"] = ("This workbook tracks whether loan proceeds have been utilised for the sanctioned "
                "purpose, in line with RBI's Fair Practices Code / KYC Master Direction expectations "
                "on monitoring end use of funds for secured retail loans (Loan Against Property and "
                "Home Loan). It should be updated by the Credit/Monitoring team after each disbursement "
                "tranche and at each periodic review (typically 30/60/90 days post-disbursement, and "
                "at each construction-linked tranche for Home Loans).")
    ws["A4"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A4:F8")

    rows = [
        ["Sheet", "What it captures", "Who updates it"],
        ["Loan Master", "Borrower, property/asset and sanction details", "Credit / Ops team, at sanction"],
        ["End-Use Declaration", "Purpose-wise break-up of sanctioned amount as declared by borrower", "RM / Credit team, at sanction"],
        ["Disbursement Register", "Tranche-wise disbursement — date, amount, mode, payee", "Disbursement / Ops team"],
        ["Txn Register (Auto)", "Every post-disbursement bank transaction, auto-classified", "Auto-generated by the Python script"],
        ["Utilisation Summary", "Category-wise utilisation vs sanctioned end-use, with variance", "Auto-generated (formulas) — review manually"],
        ["Diversion Red-Flag Log", "Transactions breaching diversion-risk rules, for investigation", "Monitoring team — mandatory sign-off"],
        ["Site Visit & Verification", "Physical verification checklist (property stage / business use)", "Field / Credit team"],
        ["Utilisation Certificate", "Final certificate for the credit file / auditor / RBI inspection", "Credit Head — signs off"],
        ["Dashboard", "One-page summary of utilisation %, flags and status", "Auto-generated"],
    ]
    r = 10
    for i, row in enumerate(rows):
        for j, val in enumerate(row):
            c = ws.cell(row=r + i, column=1 + j, value=val)
            c.font = HEADER_FONT if i == 0 else NORMAL_FONT
            c.fill = HEADER_FILL if i == 0 else (STRIPE_FILL if i % 2 == 0 else PatternFill(fill_type=None))
            c.border = BORDER
            c.alignment = WRAP
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 32

    r += len(rows) + 2
    ws.cell(row=r, column=1, value="Legend").font = SUBHEADER_FONT
    r += 1
    legend = [
        (YELLOW, "Yellow fill — cell you should fill in / edit manually"),
        (RED_FILL, "Red fill — red-flagged / not permitted for declared end use"),
        (GREEN_FILL, "Green fill — verified / permitted / within tolerance"),
    ]
    for color, text in legend:
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=color)
        ws.cell(row=r, column=2, value=text).font = NORMAL_FONT
        r += 1
    ws.sheet_view.showGridLines = False

    # =====================================================================
    # SHEET 2: LOAN MASTER
    # =====================================================================
    ws = wb.create_sheet("Loan Master")
    style_title(ws, "A1:D1", "LOAN MASTER DATA")
    ws.sheet_view.showGridLines = False

    fields = [
        ("Borrower Name", borrower_name),
        ("Loan / Account Number", loan_account_no),
        ("Loan Type", loan_type),
        ("Sanctioned Amount (Rs.)", loan_amount),
        ("Sanction Date", sanction_date),
        ("Disbursement Date (auto-detected)", disb_row["sDate"].date() if disb_row is not None else None),
        ("Disbursed Amount (auto-detected)", float(disb_row["sAmount"]) if disb_row is not None else None),
        ("Lender / NBFC", lender_hint or ""),
        ("Declared Purpose", purpose_declared),
        ("Property / Asset Address (LAP collateral or HL property)", property_address),
        ("Monitoring Window (days)", window_days),
        ("Monitoring Window End Date", summary["window_end"].date()),
        ("Prepared By", ""),
        ("Date Prepared", datetime.today().date()),
    ]
    r = 3
    for label, val in fields:
        ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
        ws.cell(row=r, column=1).border = BORDER
        c = ws.cell(row=r, column=2, value=val)
        c.font = INPUT_FONT
        c.fill = INPUT_FILL
        c.border = BORDER
        if "Amount" in label or "Disbursed" in label:
            c.number_format = '#,##0.00'
        if "Date" in label:
            c.number_format = 'dd-mmm-yyyy'
        r += 1
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 34
    ws.cell(row=r + 1, column=1, value="Note: Rows filled yellow are inputs — edit for each loan file. "
            "Disbursement date/amount are auto-detected from the bank statement by the Python script "
            "(engine matches the credit closest to the sanctioned amount).").font = NOTE_FONT
    ws.merge_cells(f"A{r+1}:D{r+1}")
    ws[f"A{r+1}"].alignment = WRAP

    # =====================================================================
    # SHEET 3: END-USE DECLARATION
    # =====================================================================
    ws = wb.create_sheet("End-Use Declaration")
    style_title(ws, "A1:E1", "SANCTIONED END-USE DECLARATION")
    ws.sheet_view.showGridLines = False
    ws["A3"] = ("Purpose-wise break-up of the sanctioned amount, as declared by the borrower at "
                "login/sanction. This is the baseline the actual utilisation (Utilisation Summary "
                "sheet) is compared against.")
    ws.merge_cells("A3:E3")
    ws["A3"].alignment = WRAP
    ws["A3"].font = NOTE_FONT

    if loan_type.upper() == "HL":
        example_rows = [
            ["Purchase consideration to seller/builder", 0.85, "", "Direct to seller/builder account only"],
            ["Stamp duty & registration", 0.06, "", ""],
            ["Interior / fit-out", 0.05, "", "Disbursed only against invoice/stage completion"],
            ["Processing fee / insurance financed", 0.02, "", ""],
            ["Contingency / buffer", 0.02, "", ""],
        ]
    else:
        example_rows = [
            ["Business working capital", 0.50, "", "GST returns / stock statement to be obtained"],
            ["Business expansion / capex", 0.25, "", "Invoice-backed disbursement preferred"],
            ["Existing high-cost debt consolidation", 0.15, "", "Closure proof from other lender required"],
            ["Working capital buffer / contingency", 0.10, "", ""],
        ]
    headers = ["Declared End-Use Head", "% of Sanctioned Amount", "Amount (Rs.)", "Supporting Document Required"]
    r = 5
    for j, h in enumerate(headers):
        ws.cell(row=r, column=1 + j, value=h)
    style_header_row(ws, r, len(headers))
    for i, row in enumerate(example_rows):
        rr = r + 1 + i
        ws.cell(row=rr, column=1, value=row[0]).font = INPUT_FONT
        ws.cell(row=rr, column=1).fill = INPUT_FILL
        pc = ws.cell(row=rr, column=2, value=row[1])
        pc.number_format = "0%"
        pc.font = INPUT_FONT
        pc.fill = INPUT_FILL
        amt = ws.cell(row=rr, column=3, value=f"=B{rr}*'Loan Master'!$B$6")
        amt.number_format = '#,##0.00'
        ws.cell(row=rr, column=4, value=row[3]).font = NORMAL_FONT
        for c in range(1, 5):
            ws.cell(row=rr, column=c).border = BORDER
            ws.cell(row=rr, column=c).alignment = WRAP
    total_row = r + 1 + len(example_rows)
    ws.cell(row=total_row, column=1, value="TOTAL").font = SUBHEADER_FONT
    ws.cell(row=total_row, column=2, value=f"=SUM(B{r+1}:B{total_row-1})").number_format = "0%"
    ws.cell(row=total_row, column=3, value=f"=SUM(C{r+1}:C{total_row-1})").number_format = '#,##0.00'
    for c in range(1, 5):
        ws.cell(row=total_row, column=c).fill = SUBHEADER_FILL
        ws.cell(row=total_row, column=c).border = BORDER
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 45

    # =====================================================================
    # SHEET 4: DISBURSEMENT REGISTER
    # =====================================================================
    ws = wb.create_sheet("Disbursement Register")
    style_title(ws, "A1:H1", "TRANCHE-WISE DISBURSEMENT REGISTER")
    ws.sheet_view.showGridLines = False
    headers = ["Tranche #", "Disbursement Date", "Amount (Rs.)", "Mode", "Paid To (Beneficiary)",
               "Beneficiary Type", "Linked Milestone / Invoice Ref.", "Cumulative % of Sanction"]
    r = 3
    for j, h in enumerate(headers):
        ws.cell(row=r, column=1 + j, value=h)
    style_header_row(ws, r, len(headers))

    disb_amt = float(disb_row["sAmount"]) if disb_row is not None else loan_amount
    beneficiary = "Builder / Seller a/c" if loan_type.upper() == "HL" else "Borrower operative a/c"
    data_row = [1, disb_row["sDate"].date() if disb_row is not None else None, disb_amt,
                str(disb_row["sMode"]) if disb_row is not None else "", beneficiary,
                "Direct / Third-Party", "Auto-detected from bank statement — verify against sanction letter", None]
    rr = r + 1
    for j, val in enumerate(data_row):
        c = ws.cell(row=rr, column=1 + j, value=val)
        c.font = NORMAL_FONT
        c.border = BORDER
        c.alignment = WRAP
    ws.cell(row=rr, column=3).number_format = '#,##0.00'
    ws.cell(row=rr, column=2).number_format = 'dd-mmm-yyyy'
    ws.cell(row=rr, column=8, value=f"=C{rr}/'Loan Master'!$B$6").number_format = "0.0%"

    # a couple of blank rows for future tranches
    for extra in range(3):
        rr += 1
        for j in range(len(headers)):
            ws.cell(row=rr, column=1 + j).border = BORDER
        ws.cell(row=rr, column=3).number_format = '#,##0.00'
        ws.cell(row=rr, column=2).number_format = 'dd-mmm-yyyy'
        ws.cell(row=rr, column=8, value=f"=SUM($C${r+1}:C{rr})/'Loan Master'!$B$6").number_format = "0.0%"
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 42
    ws.column_dimensions["H"].width = 18

    # =====================================================================
    # SHEET 5: TXN REGISTER (AUTO)
    # =====================================================================
    ws = wb.create_sheet("Txn Register (Auto)")
    style_title(ws, "A1:L1", "POST-DISBURSEMENT TRANSACTION REGISTER (AUTO-CLASSIFIED)")
    ws.sheet_view.showGridLines = False
    ws["A3"] = ("Generated by end_use_engine.py from the borrower's bank statement. Every debit/credit "
                "from the disbursement date onward is classified into an End-Use Bucket and checked "
                "against automated diversion-risk rules. Do not hand-edit — rerun the script instead.")
    ws.merge_cells("A3:L3")
    ws["A3"].font = NOTE_FONT
    ws["A3"].alignment = WRAP

    post = df[df["Days_From_Disbursement"] >= 0].copy()
    post_display = post[["sDate", "sMode", "sAmount", "sCreditOrDebit", "sNarration", "counter_party",
                          "sCategory", "End_Use_Bucket", "Permitted_For_Loan_Type", "Base_Risk_Level",
                          "Days_From_Disbursement", "Red_Flag", "Flag_Reason"]].copy()
    post_display.columns = ["Date", "Mode", "Amount", "Cr/Dr", "Narration", "Counterparty",
                             "Bank Category", "End-Use Bucket", "Permitted?", "Risk Level",
                             "Days from Disbursement", "Red Flag", "Flag Reason"]
    next_row = write_df(ws, post_display, 5, currency_cols=["Amount"], date_cols=["Date"])
    txn_header_row = 5
    txn_first_data_row = 6
    txn_last_data_row = next_row - 1

    # Conditional formatting: highlight red-flagged rows
    red_flag_col = "L"  # "Red Flag" column
    rng = f"A{txn_first_data_row}:M{txn_last_data_row}"
    ws.conditional_formatting.add(
        rng,
        FormulaRule(formula=[f"${red_flag_col}{txn_first_data_row}=TRUE"],
                    fill=PatternFill("solid", fgColor=RED_FILL))
    )
    autofit_columns(ws, max_width=45)
    ws.column_dimensions["E"].width = 42  # Narration needs room
    ws.column_dimensions["M"].width = 45  # Flag reason
    ws.freeze_panes = "A6"

    # =====================================================================
    # SHEET 6: UTILISATION SUMMARY
    # =====================================================================
    ws = wb.create_sheet("Utilisation Summary")
    style_title(ws, "A1:F1", "END-USE UTILISATION SUMMARY (POST-DISBURSEMENT WINDOW)")
    ws.sheet_view.showGridLines = False

    ws["A3"] = "Disbursed Amount (Rs.)"
    ws["B3"] = "='Loan Master'!B9"
    ws["B3"].number_format = '#,##0.00'
    ws["A4"] = "Monitoring Window"
    ws["B4"] = "=TEXT('Loan Master'!B8,\"dd-mmm-yyyy\")&\" to \"&TEXT('Loan Master'!B14,\"dd-mmm-yyyy\")"
    for cell in ["A3", "A4"]:
        ws[cell].font = SUBHEADER_FONT

    headers = ["End-Use Bucket", "Total Amount (Rs.)", "Txn Count", "% of Disbursed Amount", "Assessment"]
    r = 6
    for j, h in enumerate(headers):
        ws.cell(row=r, column=1 + j, value=h)
    style_header_row(ws, r, len(headers))

    buckets = sorted(END_USE_BUCKETS := df["End_Use_Bucket"].unique().tolist())
    rr = r
    for bucket in buckets:
        rr += 1
        ws.cell(row=rr, column=1, value=bucket).border = BORDER
        amt_formula = (f"=SUMIFS('Txn Register (Auto)'!$C:$C,'Txn Register (Auto)'!$H:$H,\"{bucket}\","
                        f"'Txn Register (Auto)'!$D:$D,\"DEBIT\")")
        ws.cell(row=rr, column=2, value=amt_formula).number_format = '#,##0.00'
        cnt_formula = (f"=COUNTIFS('Txn Register (Auto)'!$H:$H,\"{bucket}\","
                        f"'Txn Register (Auto)'!$D:$D,\"DEBIT\")")
        ws.cell(row=rr, column=3, value=cnt_formula)
        ws.cell(row=rr, column=4, value=f"=B{rr}/$B$3").number_format = "0.0%"
        assess_formula = (f'=IF(B{rr}=0,"-",IF(D{rr}>0.1,"Review — material outflow","OK"))')
        ws.cell(row=rr, column=5, value=assess_formula)
        for c in range(1, 6):
            ws.cell(row=rr, column=c).border = BORDER
            ws.cell(row=rr, column=c).font = NORMAL_FONT
        if (rr - r) % 2 == 0:
            for c in range(1, 6):
                ws.cell(row=rr, column=c).fill = STRIPE_FILL

    total_row = rr + 1
    ws.cell(row=total_row, column=1, value="TOTAL DEBITED").font = SUBHEADER_FONT
    ws.cell(row=total_row, column=2, value=f"=SUM(B{r+1}:B{rr})").number_format = '#,##0.00'
    ws.cell(row=total_row, column=3, value=f"=SUM(C{r+1}:C{rr})")
    ws.cell(row=total_row, column=4, value=f"=B{total_row}/$B$3").number_format = "0.0%"
    for c in range(1, 6):
        ws.cell(row=total_row, column=c).fill = SUBHEADER_FILL
        ws.cell(row=total_row, column=c).border = BORDER

    # Conditional formatting for Assessment column
    ws.conditional_formatting.add(
        f"E{r+1}:E{rr}",
        CellIsRule(operator="equal", formula=['"OK"'], fill=PatternFill("solid", fgColor=GREEN_FILL))
    )
    ws.conditional_formatting.add(
        f"E{r+1}:E{rr}",
        FormulaRule(formula=[f'LEFT(E{r+1},6)="Review"'], fill=PatternFill("solid", fgColor=RED_FILL))
    )
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 26

    note_row = total_row + 2
    ws.cell(row=note_row, column=1,
            value=("Note: 'Third-Party Transfer (Unverified)' and 'Debt Servicing' buckets are marked "
                   "'Review' by default — obtain invoices/purpose narration from the borrower before "
                   "closing the file; they are not automatically treated as diversion.")).font = NOTE_FONT
    ws.merge_cells(f"A{note_row}:E{note_row}")
    ws[f"A{note_row}"].alignment = WRAP

    # =====================================================================
    # SHEET 7: DIVERSION RED-FLAG LOG
    # =====================================================================
    ws = wb.create_sheet("Diversion Red-Flag Log")
    style_title(ws, "A1:I1", "DIVERSION / RED-FLAG REGISTER — REQUIRES INVESTIGATION & SIGN-OFF")
    ws.sheet_view.showGridLines = False

    flagged = post[post["Red_Flag"]][["sDate", "sAmount", "sCreditOrDebit", "sNarration",
                                       "counter_party", "Flag_Reason"]].copy()
    flagged.columns = ["Date", "Amount", "Cr/Dr", "Narration", "Counterparty", "Flag Reason"]
    flagged["Investigated By"] = ""
    flagged["Borrower Explanation"] = ""
    flagged["Status"] = "Open"

    if flagged.empty:
        ws["A3"] = "No transactions breached the automated diversion-risk rules in the monitoring window."
        ws["A3"].font = NORMAL_FONT
        next_row2 = 4
    else:
        next_row2 = write_df(ws, flagged, 3, currency_cols=["Amount"], date_cols=["Date"])
        status_col = flagged.columns.get_loc("Status") + 1
        col_letter = get_column_letter(status_col)
        dv = DataValidation(type="list", formula1='"Open,Under Review,Cleared,Escalated"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}4:{col_letter}{next_row2-1}")
        ws.conditional_formatting.add(
            f"{col_letter}4:{col_letter}{next_row2-1}",
            CellIsRule(operator="equal", formula=['"Cleared"'], fill=PatternFill("solid", fgColor=GREEN_FILL))
        )
        ws.conditional_formatting.add(
            f"{col_letter}4:{col_letter}{next_row2-1}",
            CellIsRule(operator="equal", formula=['"Escalated"'], fill=PatternFill("solid", fgColor=RED_FILL))
        )
    autofit_columns(ws, max_width=42)
    ws.column_dimensions["D"].width = 42
    ws.column_dimensions["F"].width = 45
    ws.column_dimensions["H"].width = 30

    # =====================================================================
    # SHEET 8: SITE VISIT & PHYSICAL VERIFICATION
    # =====================================================================
    ws = wb.create_sheet("Site Visit & Verification")
    style_title(ws, "A1:F1", "SITE VISIT / PHYSICAL VERIFICATION CHECKLIST")
    ws.sheet_view.showGridLines = False
    if loan_type.upper() == "HL":
        checklist = [
            "Property identified matches the one financed (address tallies with sanction letter)",
            "Stage of construction / possession matches the disbursement tranche claimed",
            "Registered sale deed / allotment letter obtained and verified",
            "No unauthorised construction / deviation from approved plan observed",
            "Photographs of property taken and filed (dated, geo-tagged if possible)",
            "Occupancy / possession certificate obtained (on completion)",
            "Property insurance in force and lender's interest noted",
        ]
    else:
        checklist = [
            "Mortgaged property physically verified — exists, matches title documents",
            "Declared business/use of funds is operational at the stated address",
            "Business registration / GST / trade license verified where applicable",
            "No signs the property or business is non-existent / shell in nature",
            "Photographs of property/business premises taken and filed",
            "Utilisation broadly consistent with the declared end-use (per Utilisation Summary)",
            "No adverse market intelligence on borrower / co-obligors",
        ]
    headers = ["#", "Verification Point", "Verified (Y/N)", "Verified By", "Date", "Remarks"]
    r = 3
    for j, h in enumerate(headers):
        ws.cell(row=r, column=1 + j, value=h)
    style_header_row(ws, r, len(headers))
    dv_yn = DataValidation(type="list", formula1='"Y,N,NA"', allow_blank=True)
    ws.add_data_validation(dv_yn)
    for i, item in enumerate(checklist):
        rr = r + 1 + i
        ws.cell(row=rr, column=1, value=i + 1)
        ws.cell(row=rr, column=2, value=item)
        vcell = ws.cell(row=rr, column=3)
        dv_yn.add(vcell.coordinate)
        ws.cell(row=rr, column=5).number_format = "dd-mmm-yyyy"
        for c in range(1, 7):
            ws.cell(row=rr, column=c).border = BORDER
            ws.cell(row=rr, column=c).font = NORMAL_FONT
            ws.cell(row=rr, column=c).alignment = WRAP
    ws.conditional_formatting.add(
        f"C{r+1}:C{r+len(checklist)}",
        CellIsRule(operator="equal", formula=['"N"'], fill=PatternFill("solid", fgColor=RED_FILL))
    )
    ws.conditional_formatting.add(
        f"C{r+1}:C{r+len(checklist)}",
        CellIsRule(operator="equal", formula=['"Y"'], fill=PatternFill("solid", fgColor=GREEN_FILL))
    )
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 35

    # =====================================================================
    # SHEET 9: UTILISATION CERTIFICATE
    # =====================================================================
    ws = wb.create_sheet("Utilisation Certificate")
    style_title(ws, "A1:D1", "CERTIFICATE OF END-USE VERIFICATION")
    ws.sheet_view.showGridLines = False
    body = [
        "This is to certify that the post-disbursement end-use of funds review for the loan account "
        "referenced below has been carried out based on a review of the borrower's bank statement, "
        "supporting documents and (where applicable) physical/site verification.",
        "",
        "Borrower Name: ='Loan Master'!B3",
        "Loan / Account Number: ='Loan Master'!B4",
        "Sanctioned Amount: ='Loan Master'!B6",
        "Disbursed Amount: ='Loan Master'!B9",
        "Disbursement Date: ='Loan Master'!B8",
        "Declared Purpose: ='Loan Master'!B11",
        "",
        "Total Utilisation Verified (per Utilisation Summary): ='Utilisation Summary'!B" + str(total_row),
        "% of Disbursed Amount Utilised: ='Utilisation Summary'!D" + str(total_row),
        "Number of Red-Flagged Transactions: =COUNTA('Diversion Red-Flag Log'!A4:A1000)-COUNTBLANK('Diversion Red-Flag Log'!A4:A1000)",
        "",
        "Finding: Based on the above review, the utilisation of loan proceeds is assessed as "
        "[Satisfactory / Requires Further Clarification / Diversion Suspected] — strike out as applicable.",
        "",
        "Remarks:",
    ]
    r = 3
    for line in body:
        if line.startswith(("Borrower", "Loan / Account", "Sanctioned", "Disbursed", "Disbursement",
                             "Declared", "Total Utilisation", "% of Disbursed", "Number of")):
            label, _, formula = line.partition(": ")
            ws.cell(row=r, column=1, value=label + ":").font = SUBHEADER_FONT
            fc = ws.cell(row=r, column=2, value=formula)
            fc.font = NORMAL_FONT
            if "%" in label:
                fc.number_format = "0.0%"
            elif "Amount" in label:
                fc.number_format = '#,##0.00'
            elif "Date" in label:
                fc.number_format = "dd-mmm-yyyy"
        else:
            ws.cell(row=r, column=1, value=line).font = NORMAL_FONT if line else NORMAL_FONT
            ws.merge_cells(f"A{r}:D{r}")
            ws[f"A{r}"].alignment = WRAP
        r += 1
    r += 2
    ws.cell(row=r, column=1, value="Remarks Box:").font = SUBHEADER_FONT
    r += 1
    ws.merge_cells(f"A{r}:D{r+3}")
    ws.cell(row=r, column=1).border = BORDER
    ws.cell(row=r, column=1).fill = INPUT_FILL
    r += 6
    sign_labels = ["Verified By (Name & Designation):", "Signature & Date:",
                   "Approved By (Credit Head):", "Signature & Date:"]
    for lbl in sign_labels:
        ws.cell(row=r, column=1, value=lbl).font = NORMAL_FONT
        ws.cell(row=r, column=2).border = Border(bottom=Side(style="thin"))
        r += 2
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    # =====================================================================
    # SHEET 10: DASHBOARD
    # =====================================================================
    ws = wb.create_sheet("Dashboard")
    style_title(ws, "A1:D1", "END-USE MONITORING — SUMMARY DASHBOARD")
    ws.sheet_view.showGridLines = False
    kpis = [
        ("Borrower", "='Loan Master'!B3"),
        ("Loan Type", "='Loan Master'!B5"),
        ("Sanctioned Amount", "='Loan Master'!B6"),
        ("Disbursed Amount", "='Loan Master'!B9"),
        ("Disbursement Date", "='Loan Master'!B8"),
        ("Total Utilised (window)", f"='Utilisation Summary'!B{total_row}"),
        ("% Utilised", f"='Utilisation Summary'!D{total_row}"),
        ("Red-Flagged Transactions", "=COUNTIF('Txn Register (Auto)'!L:L,TRUE)"),
        ("Red-Flagged Amount", "=SUMIF('Txn Register (Auto)'!L:L,TRUE,'Txn Register (Auto)'!C:C)"),
        ("Overall Status", '=IF(COUNTIF(\'Txn Register (Auto)\'!L:L,TRUE)=0,"CLEAN","REVIEW REQUIRED")'),
    ]
    r = 3
    for label, formula in kpis:
        ws.cell(row=r, column=1, value=label).font = SUBHEADER_FONT
        ws.cell(row=r, column=1).border = BORDER
        c = ws.cell(row=r, column=2, value=formula)
        c.border = BORDER
        c.font = Font(name=FONT_NAME, size=11, bold=True)
        if "Amount" in label or "Utilised" in label and "%" not in label:
            c.number_format = '#,##0.00'
        if "%" in label:
            c.number_format = "0.0%"
        if "Date" in label:
            c.number_format = "dd-mmm-yyyy"
        r += 1
    ws.conditional_formatting.add(
        "B12", CellIsRule(operator="equal", formula=['"CLEAN"'], fill=PatternFill("solid", fgColor=GREEN_FILL))
    )
    ws.conditional_formatting.add(
        "B12", CellIsRule(operator="equal", formula=['"REVIEW REQUIRED"'], fill=PatternFill("solid", fgColor=RED_FILL))
    )
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 26

    # Bucket-wise mini chart data + bar chart
    from openpyxl.chart import BarChart, Reference
    chart_start = r + 2
    ws.cell(row=chart_start, column=1, value="Utilisation by End-Use Bucket").font = SUBHEADER_FONT
    chart_start += 1
    n_buckets = len(buckets)
    for i in range(n_buckets):
        src_row = 7 + i  # matches Utilisation Summary layout (header at row6, data from row7)
        ws.cell(row=chart_start + i, column=1, value=f"='Utilisation Summary'!A{src_row}")
        ws.cell(row=chart_start + i, column=2, value=f"='Utilisation Summary'!B{src_row}").number_format = '#,##0'
    chart = BarChart()
    chart.title = "Utilisation by End-Use Bucket (Rs.)"
    chart.type = "bar"
    chart.style = 10
    data = Reference(ws, min_col=2, min_row=chart_start, max_row=chart_start + n_buckets - 1)
    cats = Reference(ws, min_col=1, min_row=chart_start, max_row=chart_start + n_buckets - 1)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    chart.legend = None
    chart.height = 8
    chart.width = 18
    ws.add_chart(chart, f"D{3}")

    # Reorder sheets sensibly
    order = ["Instructions", "Loan Master", "End-Use Declaration", "Disbursement Register",
             "Txn Register (Auto)", "Utilisation Summary", "Diversion Red-Flag Log",
             "Site Visit & Verification", "Utilisation Certificate", "Dashboard"]
    wb._sheets = [wb[name] for name in order]
    wb.active = wb.sheetnames.index("Dashboard")

    # Print setup: landscape, fit to page width, repeat header row 5 on Txn Register
    for name in order:
        sh = wb[name]
        sh.page_setup.orientation = "landscape"
        sh.page_setup.fitToWidth = 1
        sh.page_setup.fitToHeight = 0
        sh.sheet_properties.pageSetUpPr.fitToPage = True
    wb["Txn Register (Auto)"].print_title_rows = "5:5"

    wb.save(out_path)
    return out_path, df, disb_row, summary


if __name__ == "__main__":
    out_path, df, disb_row, summary = build_workbook(
        statement_path="/mnt/user-data/uploads/Book8.xlsx",
        loan_amount=892242,
        loan_type="LAP",
        lender_hint="SBFC",
        borrower_name="[Borrower Name — from KYC]",
        loan_account_no="[LAN — Loan Account Number]",
        sanction_date=None,
        purpose_declared="Business expansion / working capital (LAP)",
        property_address="[Mortgaged property address]",
        out_path="/home/claude/work/EUF_Monitoring_LAP_HL.xlsx",
        window_days=90,
    )
    print("Saved:", out_path)
