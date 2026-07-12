"""
batch_report.py
----------------
Scan a FOLDER of bank-statement Excel files (700 of them, in your case) and
produce ONE consolidated "Portfolio End-Use of Funds Report.xlsx" covering
every loan account, without having to open each file by hand.

WORKS IN 3 MODES (auto-detected from what you give it):
  1. With a loan master mapping (recommended) -- a CSV/XLSX listing each
     filename with its real loan_amount / loan_type / borrower etc. Every
     matched file uses the REAL sanctioned amount to find the disbursement
     and compute accurate % utilised.
  2. Without a mapping -- for any file not found in the mapping (or if you
     don't have one at all), the script auto-detects the disbursement as
     the single largest CREDIT in that statement. Less precise, but works
     out of the box on all 700 files today.
  3. Mixed -- most common in practice: some files matched via the mapping,
     the rest auto-detected. Every row in the output says which mode was
     used ("Matched" / "Auto-detected"), so you can see which numbers to
     trust more.

USAGE
-----
    pip install openpyxl pandas --break-system-packages

    # No mapping file -- auto-detect everything
    python3 batch_report.py --folder "C:/Loans/Statements" --out portfolio_report.xlsx

    # With a mapping file (columns: filename, loan_amount, loan_type,
    # lender_hint, borrower_name, loan_account_no, sanction_date,
    # purpose_declared -- only 'filename' and 'loan_amount' are required,
    # the rest are optional)
    python3 batch_report.py --folder "C:/Loans/Statements" --mapping loan_master.csv --out portfolio_report.xlsx

The 700 statement files themselves are NOT modified or uploaded anywhere --
everything runs locally on your machine.
"""

import argparse
import os
import sys
import traceback
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.chart import BarChart, Reference

from end_use_engine import run as run_engine, END_USE_MAP

# ---------------------------------------------------------------------------
# STYLE CONSTANTS (match the single-loan workbook look)
# ---------------------------------------------------------------------------
FONT_NAME = "Arial"
NAVY = "1F3864"
LIGHT_BLUE = "DCE6F1"
RED_FILL = "FFC7CE"
GREEN_FILL = "C6EFCE"
GREY = "F2F2F2"

TITLE_FONT = Font(name=FONT_NAME, size=14, bold=True, color="FFFFFF")
HEADER_FONT = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
NORMAL_FONT = Font(name=FONT_NAME, size=10)
NOTE_FONT = Font(name=FONT_NAME, size=9, italic=True, color="808080")
TITLE_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
STRIPE_FILL = PatternFill("solid", fgColor=GREY)
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_title(ws, cell_range, text):
    ws.merge_cells(cell_range)
    c = ws[cell_range.split(":")[0]]
    c.value = text
    c.font = TITLE_FONT
    c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[c.row].height = 26


def style_header_row(ws, row, ncols, start_col=1):
    for i in range(start_col, start_col + ncols):
        c = ws.cell(row=row, column=i)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = 32


def autofit_columns(ws, min_width=10, max_width=50, extra=2):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            length = len(str(cell.value))
            widths[cell.column_letter] = max(widths.get(cell.column_letter, 0), length)
    for col, length in widths.items():
        ws.column_dimensions[col].width = max(min_width, min(max_width, length + extra))


def load_mapping(path):
    if not path:
        return {}
    if path.lower().endswith(".csv"):
        m = pd.read_csv(path)
    else:
        m = pd.read_excel(path)
    m.columns = [c.strip().lower() for c in m.columns]
    if "filename" not in m.columns:
        raise ValueError("Mapping file must have a 'filename' column matching the statement file names.")
    return {str(row["filename"]).strip(): row.to_dict() for _, row in m.iterrows()}


def find_statement_files(folder, recursive=True):
    """
    Find all .xlsx/.xls files under `folder`.
    Handles BOTH layouts:
      - flat: folder/loan1.xlsx, folder/loan2.xlsx, ...
      - one-subfolder-per-loan (common with BSA report downloads):
        folder/SBFCLAP0000376039/report.xlsx,
        folder/SBFCLAP0000382737/report.xlsx, ...
    Returns a list of dicts: {path, rel_path, folder_name, filename}
    `folder_name` is the immediate parent directory name relative to `folder`
    itself (e.g. 'SBFCLAP0000376039') -- used as a loan-account-number guess
    when a file isn't in the mapping and its own sheets don't carry one.
    """
    results = []
    if recursive:
        for root, dirs, files in os.walk(folder):
            for f in files:
                if f.lower().endswith((".xlsx", ".xls")) and not f.startswith("~$"):
                    full = os.path.join(root, f)
                    rel = os.path.relpath(full, folder)
                    parent = os.path.basename(root)
                    results.append({
                        "path": full, "rel_path": rel,
                        "folder_name": parent if parent != os.path.basename(folder.rstrip(os.sep)) else "",
                        "filename": f,
                    })
    else:
        for f in os.listdir(folder):
            if f.lower().endswith((".xlsx", ".xls")) and not f.startswith("~$"):
                results.append({"path": os.path.join(folder, f), "rel_path": f, "folder_name": "", "filename": f})
    results.sort(key=lambda r: r["rel_path"])
    return results


def process_folder(folder, mapping_path=None, window_days=90, out_path="portfolio_report.xlsx",
                    default_loan_type="LAP", progress_every=25, recursive=True):
    mapping = load_mapping(mapping_path)
    files = find_statement_files(folder, recursive=recursive)
    print(f"Found {len(files)} Excel files under {folder} (recursive={recursive})")
    if mapping:
        print(f"Loaded mapping for {len(mapping)} file(s) from {mapping_path}")

    results = []       # one row per loan for Portfolio Summary
    all_flags = []      # consolidated red-flag rows across every file
    bucket_totals = {}  # portfolio-wide End-Use Bucket -> amount
    errors = []          # files that failed to process

    for i, entry in enumerate(files, 1):
        fpath = entry["path"]
        rel_name = entry["rel_path"]                 # used as the unique "File" identifier
        # Mapping can be keyed by the plain filename OR the folder name (e.g. the LAN),
        # whichever matches -- try filename first, then folder name.
        row_map = mapping.get(entry["filename"]) or mapping.get(entry["folder_name"]) or {}
        loan_amount = row_map.get("loan_amount") or None
        loan_type = str(row_map.get("loan_type") or default_loan_type).upper()
        lender_hint = row_map.get("lender_hint") or None
        borrower_name = row_map.get("borrower_name") or ""
        loan_account_no = row_map.get("loan_account_no") or entry["folder_name"] or ""
        matched = bool(row_map)

        try:
            df, disb_row, summary = run_engine(fpath, loan_amount=loan_amount, loan_type=loan_type,
                                                lender_hint=lender_hint, window_days=window_days)
            if disb_row is None:
                raise ValueError("No CREDIT transactions found — cannot identify a disbursement.")

            effective_amount = float(loan_amount) if loan_amount else float(disb_row["sAmount"])
            n_flags = int(df["Red_Flag"].sum())
            flagged_amount = float(df.loc[df["Red_Flag"], "sAmount"].sum())
            top_bucket_row = summary["by_bucket"].iloc[0] if not summary["by_bucket"].empty else None

            results.append({
                "File": rel_name,
                "Borrower": borrower_name,
                "Loan A/C No.": loan_account_no,
                "Loan Type": loan_type,
                "Amount Source": "Matched (mapping)" if matched else "Auto-detected",
                "Disbursement Date": disb_row["sDate"].date(),
                "Disbursed / Sanctioned Amount": effective_amount,
                "Total Utilised (window)": summary["total_utilised_in_window"],
                "% Utilised": summary["pct_utilised"],
                "Top End-Use Bucket": top_bucket_row["End_Use_Bucket"] if top_bucket_row is not None else "",
                "Top Bucket %": (top_bucket_row["Pct_of_Loan_Amount"] if top_bucket_row is not None else None),
                "Red Flag Count": n_flags,
                "Red Flag Amount": flagged_amount,
                "Status": "REVIEW REQUIRED" if n_flags > 0 else "CLEAN",
            })

            for _, r in df[df["Red_Flag"]].iterrows():
                all_flags.append({
                    "File": rel_name, "Borrower": borrower_name, "Date": r["sDate"].date(),
                    "Amount": r["sAmount"], "Narration": r["sNarration"],
                    "Counterparty": r.get("counter_party", ""), "Flag Reason": r["Flag_Reason"],
                })

            post = df[df["Days_From_Disbursement"] >= 0]
            post = post[post["sCreditOrDebit"].str.upper() == "DEBIT"]
            for bucket, amt in post.groupby("End_Use_Bucket")["sAmount"].sum().items():
                bucket_totals[bucket] = bucket_totals.get(bucket, 0) + amt

        except Exception as e:
            errors.append({"File": rel_name, "Error": str(e)})

        if i % progress_every == 0 or i == len(files):
            print(f"  processed {i}/{len(files)}  (errors so far: {len(errors)})")

    write_report(results, all_flags, bucket_totals, errors, out_path, window_days, len(files))
    print(f"\nDone. Report saved to: {out_path}")
    print(f"  Loans processed OK : {len(results)}")
    print(f"  Files with errors  : {len(errors)}")
    print(f"  Loans flagged      : {sum(1 for r in results if r['Status'] == 'REVIEW REQUIRED')}")
    return out_path


def write_df_sheet(ws, headers, rows, start_row, currency_cols=None, pct_cols=None, date_cols=None):
    currency_cols = currency_cols or []
    pct_cols = pct_cols or []
    date_cols = date_cols or []
    for j, h in enumerate(headers):
        ws.cell(row=start_row, column=1 + j, value=h)
    style_header_row(ws, start_row, len(headers))
    for i, row in enumerate(rows):
        rr = start_row + 1 + i
        for j, h in enumerate(headers):
            c = ws.cell(row=rr, column=1 + j, value=row.get(h))
            c.font = NORMAL_FONT
            c.border = BORDER
            c.alignment = WRAP
            if h in currency_cols:
                c.number_format = '#,##0.00'
            if h in pct_cols:
                c.number_format = '0.0%'
            if h in date_cols:
                c.number_format = 'dd-mmm-yyyy'
        if i % 2 == 1:
            for j in range(len(headers)):
                ws.cell(row=rr, column=1 + j).fill = STRIPE_FILL
    return start_row + 1 + len(rows)


def write_report(results, all_flags, bucket_totals, errors, out_path, window_days, n_files):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ---------------- DASHBOARD ----------------
    ws = wb.create_sheet("Dashboard")
    style_title(ws, "A1:D1", "PORTFOLIO END-USE OF FUNDS — SUMMARY DASHBOARD")
    ws.sheet_view.showGridLines = False
    n_ok = len(results)
    n_flagged = sum(1 for r in results if r["Status"] == "REVIEW REQUIRED")
    total_disbursed = sum(r["Disbursed / Sanctioned Amount"] for r in results)
    total_flag_amt = sum(r["Red Flag Amount"] for r in results)
    kpis = [
        ("Files Scanned", n_files),
        ("Loans Processed Successfully", n_ok),
        ("Files with Errors (see 'Errors' sheet)", len(errors)),
        ("Loans Flagged — Review Required", n_flagged),
        ("Loans Clean", n_ok - n_flagged),
        ("Total Disbursed / Sanctioned (Rs.)", total_disbursed),
        ("Total Red-Flagged Amount (Rs.)", total_flag_amt),
        ("Monitoring Window (days)", window_days),
        ("Report Generated", datetime.today().strftime("%d-%b-%Y %H:%M")),
    ]
    r = 3
    for label, val in kpis:
        ws.cell(row=r, column=1, value=label).font = Font(name=FONT_NAME, bold=True, color=NAVY)
        ws.cell(row=r, column=1).border = BORDER
        c = ws.cell(row=r, column=2, value=val)
        c.border = BORDER
        c.font = Font(name=FONT_NAME, size=11, bold=True)
        if "Rs." in label:
            c.number_format = '#,##0.00'
        r += 1
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 24

    # bucket chart data
    chart_row = r + 2
    ws.cell(row=chart_row, column=1, value="Portfolio Utilisation by End-Use Bucket").font = Font(bold=True, color=NAVY)
    chart_row += 1
    sorted_buckets = sorted(bucket_totals.items(), key=lambda x: -x[1])
    for i, (bucket, amt) in enumerate(sorted_buckets):
        ws.cell(row=chart_row + i, column=1, value=bucket)
        ws.cell(row=chart_row + i, column=2, value=amt).number_format = '#,##0'
    if sorted_buckets:
        chart = BarChart()
        chart.title = "Portfolio-Wide Utilisation by End-Use Bucket (Rs.)"
        chart.type = "bar"
        chart.style = 10
        data = Reference(ws, min_col=2, min_row=chart_row, max_row=chart_row + len(sorted_buckets) - 1)
        cats = Reference(ws, min_col=1, min_row=chart_row, max_row=chart_row + len(sorted_buckets) - 1)
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(cats)
        chart.legend = None
        chart.height = 9
        chart.width = 20
        ws.add_chart(chart, "D3")

    # ---------------- PORTFOLIO SUMMARY ----------------
    ws = wb.create_sheet("Portfolio Summary")
    style_title(ws, "A1:M1", "PORTFOLIO SUMMARY — ONE ROW PER LOAN")
    ws.sheet_view.showGridLines = False
    headers = ["File", "Borrower", "Loan A/C No.", "Loan Type", "Amount Source", "Disbursement Date",
               "Disbursed / Sanctioned Amount", "Total Utilised (window)", "% Utilised",
               "Top End-Use Bucket", "Top Bucket %", "Red Flag Count", "Red Flag Amount", "Status"]
    last_row = write_df_sheet(ws, headers, results, 3,
                               currency_cols=["Disbursed / Sanctioned Amount", "Total Utilised (window)", "Red Flag Amount"],
                               pct_cols=["% Utilised", "Top Bucket %"], date_cols=["Disbursement Date"])
    status_col_idx = headers.index("Status") + 1
    status_col = openpyxl.utils.get_column_letter(status_col_idx)
    ws.conditional_formatting.add(
        f"{status_col}4:{status_col}{last_row-1}",
        CellIsRule(operator="equal", formula=['"CLEAN"'], fill=PatternFill("solid", fgColor=GREEN_FILL)))
    ws.conditional_formatting.add(
        f"{status_col}4:{status_col}{last_row-1}",
        CellIsRule(operator="equal", formula=['"REVIEW REQUIRED"'], fill=PatternFill("solid", fgColor=RED_FILL)))
    autofit_columns(ws)
    ws.freeze_panes = "A4"

    # ---------------- ALL RED FLAGS ----------------
    ws = wb.create_sheet("All Red Flags")
    style_title(ws, "A1:G1", "CONSOLIDATED RED-FLAG REGISTER — ACROSS ALL FILES")
    ws.sheet_view.showGridLines = False
    if all_flags:
        headers = ["File", "Borrower", "Date", "Amount", "Narration", "Counterparty", "Flag Reason"]
        write_df_sheet(ws, headers, all_flags, 3, currency_cols=["Amount"], date_cols=["Date"])
    else:
        ws["A3"] = "No red-flagged transactions found across the portfolio."
        ws["A3"].font = NORMAL_FONT
    autofit_columns(ws)
    ws.column_dimensions["E"].width = 45
    ws.column_dimensions["G"].width = 45
    ws.freeze_panes = "A4"

    # ---------------- CATEGORY AGGREGATE ----------------
    ws = wb.create_sheet("Category Aggregate")
    style_title(ws, "A1:C1", "PORTFOLIO-WIDE UTILISATION BY END-USE BUCKET")
    ws.sheet_view.showGridLines = False
    total_all = sum(bucket_totals.values()) or 1
    rows = [{"End-Use Bucket": b, "Total Amount (Rs.)": a, "% of Portfolio Debits": a / total_all}
            for b, a in sorted_buckets]
    write_df_sheet(ws, ["End-Use Bucket", "Total Amount (Rs.)", "% of Portfolio Debits"], rows, 3,
                   currency_cols=["Total Amount (Rs.)"], pct_cols=["% of Portfolio Debits"])
    autofit_columns(ws)

    # ---------------- ERRORS ----------------
    ws = wb.create_sheet("Errors")
    style_title(ws, "A1:B1", "FILES THAT COULD NOT BE PROCESSED")
    ws.sheet_view.showGridLines = False
    if errors:
        write_df_sheet(ws, ["File", "Error"], errors, 3)
    else:
        ws["A3"] = "No errors — all files processed successfully."
        ws["A3"].font = NORMAL_FONT
    autofit_columns(ws)
    ws.column_dimensions["B"].width = 60

    order = ["Dashboard", "Portfolio Summary", "All Red Flags", "Category Aggregate", "Errors"]
    wb._sheets = [wb[name] for name in order]
    wb.active = 0
    for name in order:
        wb[name].page_setup.orientation = "landscape"
        wb[name].page_setup.fitToWidth = 1
        wb[name].page_setup.fitToHeight = 0
        wb[name].sheet_properties.pageSetUpPr.fitToPage = True

    wb.save(out_path)


if __name__ == "__main__":
    p = argparse.ArgumentParser(description="Batch End-Use-of-Funds report over a folder of bank statements")
    p.add_argument("--folder", required=True, help="Top folder containing the statement files — searched recursively by default, so one-subfolder-per-loan layouts (e.g. folder/LAN12345/report.xlsx) work out of the box")
    p.add_argument("--mapping", default=None, help="Optional CSV/XLSX with columns: filename, loan_amount, loan_type, lender_hint, borrower_name, loan_account_no. 'filename' can match either the file's own name OR its parent folder name (e.g. the LAN)")
    p.add_argument("--out", default="portfolio_report.xlsx")
    p.add_argument("--loan-type", default="LAP", choices=["LAP", "HL"], help="Default loan type for files not in the mapping")
    p.add_argument("--window-days", type=int, default=90)
    p.add_argument("--no-recursive", action="store_true", help="Only scan the top-level folder, don't look inside subfolders")
    args = p.parse_args()

    process_folder(args.folder, mapping_path=args.mapping, window_days=args.window_days,
                    out_path=args.out, default_loan_type=args.loan_type, recursive=not args.no_recursive)
