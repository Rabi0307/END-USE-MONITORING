"""
build_from_bsa_report.py
--------------------------
End-to-end EUF one-pager builder for a FULL Bank Statement Analyzer (BSA)
report (the multi-sheet format with 'Customer & Account Information',
'Overall Analysis', 'Fixed Obligations', 'AllAccountXns', etc. -- e.g.
report3_...xlsx).

Unlike build_onepager.py (which needs you to type in the borrower name,
account number etc. by hand), this script pulls all of that straight out
of the report itself, and adds an extra "Existing Obligations" panel using
the report's own Fixed Obligations / Overall Analysis sheets -- so you can
see at a glance whether the borrower already has other EMIs running and
what their FOIR looks like, alongside the usual end-use classification.

USAGE
    python3 build_from_bsa_report.py --input report3_....xlsx \
        --loan-amount 892242 --loan-type LAP --lender-hint SBFC --out onepager.xlsx

If --loan-amount is omitted, the disbursement is auto-detected as the
largest CREDIT in the statement.
"""
import argparse

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule

from end_use_engine import run as run_engine
from bsa_report_reader import read_bsa_report, summarize_existing_obligations

FONT_NAME = "Arial"
NAVY = "1F3864"
LIGHT_BLUE = "DCE6F1"
YELLOW = "FFF2CC"
RED_FILL = "FFC7CE"
GREEN_FILL = "C6EFCE"
ORANGE_FILL = "FCE4D6"
GREY = "F2F2F2"

TITLE_FONT = Font(name=FONT_NAME, size=13, bold=True, color="FFFFFF")
SECTION_FONT = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
LABEL_FONT = Font(name=FONT_NAME, size=9, bold=True, color=NAVY)
VALUE_FONT = Font(name=FONT_NAME, size=9)
TITLE_FILL = PatternFill("solid", fgColor=NAVY)
SECTION_FILL = PatternFill("solid", fgColor=NAVY)
INPUT_FILL = PatternFill("solid", fgColor=YELLOW)
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")


def section_header(ws, cell_range, text, fill=SECTION_FILL):
    ws.merge_cells(cell_range)
    c = ws[cell_range.split(":")[0]]
    c.value = text
    c.font = SECTION_FONT
    c.fill = fill
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[c.row].height = 16


def label_value(ws, label_cell, value_cell, label, value, number_format=None, fill=None):
    ws[label_cell] = label
    ws[label_cell].font = LABEL_FONT
    ws[value_cell] = value
    ws[value_cell].font = VALUE_FONT
    if fill:
        ws[value_cell].fill = PatternFill("solid", fgColor=fill)
    elif value in ("", None):
        ws[value_cell].fill = PatternFill("solid", fgColor=YELLOW)
    if number_format:
        ws[value_cell].number_format = number_format
    for cell in (label_cell, value_cell):
        ws[cell].border = BORDER


def build(input_path, loan_amount, loan_type, lender_hint, loan_account_no,
          purpose_declared, out_path, window_days=90):

    bsa = read_bsa_report(input_path)
    obligations = summarize_existing_obligations(bsa)
    df, disb_row, summary = run_engine(input_path, loan_amount=loan_amount,
                                        loan_type=loan_type, lender_hint=lender_hint,
                                        window_days=window_days)
    disb_amt = float(disb_row["sAmount"]) if disb_row is not None else loan_amount
    n_flags = int(df["Red_Flag"].sum())
    flagged_amt = float(df.loc[df["Red_Flag"], "sAmount"].sum())
    top_flags = df[df["Red_Flag"]].sort_values("sAmount", ascending=False).head(5)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EUF Summary"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    ws["A1"] = "POST-DISBURSEMENT END-USE OF FUNDS — SUMMARY (from BSA Report)"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 22

    # ---- LOAN / BORROWER DETAILS (auto-filled from the BSA report) ----
    section_header(ws, "A3:H3", "BORROWER & LOAN DETAILS  (auto-filled from report)")
    label_value(ws, "A4", "B4", "Borrower", obligations["borrower_name"], fill=GREEN_FILL)
    label_value(ws, "C4", "D4", "Account No.", obligations["account_no"], fill=GREEN_FILL)
    label_value(ws, "E4", "F4", "Bank", obligations["bank"], fill=GREEN_FILL)
    label_value(ws, "G4", "H4", "Account Type", obligations["account_type"], fill=GREEN_FILL)
    label_value(ws, "A5", "B5", "Loan A/C No. (LAN)", loan_account_no)
    label_value(ws, "C5", "D5", "Loan Type", loan_type)
    label_value(ws, "E5", "F5", "Disbursement Date",
                disb_row["sDate"].date() if disb_row is not None else None, "dd-mmm-yyyy")
    label_value(ws, "G5", "H5", "Disbursed Amount (Rs.)", disb_amt, "#,##0")
    ws.merge_cells("B6:D6")
    label_value(ws, "A6", "B6", "Declared Purpose", purpose_declared)
    label_value(ws, "E6", "F6", "Analysis Period", obligations["analysis_period"], fill=GREEN_FILL)
    label_value(ws, "G6", "H6", "Window Ends", summary["window_end"].date(), "dd-mmm-yyyy")

    # ---- EXISTING OBLIGATIONS (new — from Overall Analysis / Fixed Obligations) ----
    section_header(ws, "A8:H8", "EXISTING OBLIGATIONS ON THIS ACCOUNT  (independent of the tracked loan)")
    foir = obligations["foir"]
    foir_display = foir if foir not in (None, "inf") else "∞ (no verified income detected)"
    label_value(ws, "A9", "B9", "EMI Already Detected?", obligations["emi_detected"],
                fill=RED_FILL if str(obligations["emi_detected"]) == "True" else GREEN_FILL)
    label_value(ws, "C9", "D9", "Existing EMI Total (Rs.)",
                obligations["existing_emi_from_fixed_obligations_sheet"], "#,##0",
                fill=ORANGE_FILL if obligations["existing_emi_from_fixed_obligations_sheet"] else GREEN_FILL)
    label_value(ws, "E9", "F9", "FOIR", foir_display,
                fill=RED_FILL if foir in (None, "inf") else None)
    label_value(ws, "G9", "H9", "Cheque Bounce?", obligations["cheque_bounce"],
                fill=RED_FILL if str(obligations["cheque_bounce"]) == "True" else GREEN_FILL)
    label_value(ws, "A10", "B10", "Monthly Avg. Balance (Rs.)", obligations["monthly_avg_balance"], "#,##0")
    label_value(ws, "C10", "D10", "Current Balance (Rs.)", obligations["current_balance"], "#,##0")
    label_value(ws, "E10", "F10", "Salary Credits (Rs.)", obligations["salary_credits"], "#,##0")
    label_value(ws, "G10", "H10", "Non-Salary Income (Rs.)", obligations["non_salary_income"], "#,##0")
    ws.merge_cells("A11:H11")
    ws["A11"] = ("Note: 'Existing EMI Total' comes from the report's own Fixed Obligations sheet — "
                 "it reflects debt already being serviced on this account, separate from the loan "
                 "being tracked here. Cross-check against the 'Debt Servicing' row below: if that "
                 "figure is materially higher than this one, the tracked loan's proceeds may "
                 "themselves be funding EMIs on other debts.")
    ws["A11"].font = Font(name=FONT_NAME, size=8, italic=True, color="808080")
    ws["A11"].alignment = WRAP
    ws.row_dimensions[11].height = 26

    # ---- UTILISATION SUMMARY (left) + RED FLAG SUMMARY (right) ----
    section_header(ws, "A13:D13", "UTILISATION SUMMARY (vs. Disbursed Amount)")
    section_header(ws, "E13:H13", "RED-FLAG SUMMARY")
    headers = ["End-Use Bucket", "Amount (Rs.)", "%", "Assessment"]
    for j, h in enumerate(headers):
        c = ws.cell(row=14, column=1 + j, value=h)
        c.font = LABEL_FONT
        c.border = BORDER
        c.fill = PatternFill("solid", fgColor=LIGHT_BLUE)

    bucket_df = summary["by_bucket"].head(6)
    r = 15
    for _, row in bucket_df.iterrows():
        pct = row["Pct_of_Loan_Amount"]
        assessment = "Review" if pct > 0.10 else "OK"
        ws.cell(row=r, column=1, value=row["End_Use_Bucket"]).font = VALUE_FONT
        ws.cell(row=r, column=2, value=row["Total_Amount"]).number_format = "#,##0"
        ws.cell(row=r, column=3, value=pct).number_format = "0.0%"
        ws.cell(row=r, column=4, value=assessment)
        for c in range(1, 5):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).font = VALUE_FONT
        r += 1
    last_bucket_row = r - 1
    ws.conditional_formatting.add(f"D15:D{last_bucket_row}",
        CellIsRule(operator="equal", formula=['"OK"'], fill=PatternFill("solid", fgColor=GREEN_FILL)))
    ws.conditional_formatting.add(f"D15:D{last_bucket_row}",
        CellIsRule(operator="equal", formula=['"Review"'], fill=PatternFill("solid", fgColor=RED_FILL)))

    ws["E14"] = "Total Red-Flagged Txns"
    ws["E14"].font = LABEL_FONT
    ws["F14"] = n_flags
    ws["F14"].font = Font(name=FONT_NAME, size=11, bold=True, color="9C0006" if n_flags else "006100")
    ws["G14"] = "Flagged Amount (Rs.)"
    ws["G14"].font = LABEL_FONT
    ws["H14"] = flagged_amt
    ws["H14"].number_format = "#,##0"
    ws["H14"].font = Font(name=FONT_NAME, size=10, bold=True)
    ws.merge_cells("E15:H15")
    ws["E15"] = "Top flagged transactions:"
    ws["E15"].font = LABEL_FONT
    rr = 16
    if top_flags.empty:
        ws.merge_cells(f"E{rr}:H{rr}")
        ws[f"E{rr}"] = "None — no automated red flags in this window."
        ws[f"E{rr}"].font = VALUE_FONT
    else:
        for _, fr in top_flags.iterrows():
            ws.merge_cells(f"E{rr}:H{rr}")
            txt = f"{fr['sDate'].date()} — Rs.{fr['sAmount']:,.0f} — {fr['Flag_Reason']}"
            c = ws[f"E{rr}"]
            c.value = txt
            c.font = Font(name=FONT_NAME, size=8)
            c.fill = PatternFill("solid", fgColor=RED_FILL)
            c.alignment = WRAP
            rr += 1

    status_row = max(last_bucket_row, rr) + 1
    ws.merge_cells(f"A{status_row}:H{status_row}")
    overall = "REVIEW REQUIRED" if (n_flags > 0 or str(obligations["cheque_bounce"]) == "True") else "CLEAN"
    ws[f"A{status_row}"] = f"OVERALL STATUS: {overall}"
    ws[f"A{status_row}"].font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
    ws[f"A{status_row}"].fill = PatternFill("solid", fgColor="9C0006" if overall == "REVIEW REQUIRED" else "006100")
    ws[f"A{status_row}"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[status_row].height = 18

    link_row = status_row + 2
    ws.merge_cells(f"A{link_row}:H{link_row}")
    link_cell = ws[f"A{link_row}"]
    link_cell.value = "View full transaction detail  →  see the 'Txn Detail' tab (bottom of workbook)"
    link_cell.hyperlink = "#'Txn Detail'!A1"
    link_cell.font = Font(name=FONT_NAME, size=9, bold=True, color="0563C1", underline="single")

    widths = {"A": 22, "B": 16, "C": 16, "D": 16, "E": 22, "F": 14, "G": 20, "H": 16}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.page_margins.top = ws.page_margins.bottom = 0.4

    # ---- TRANSACTION DETAIL SHEET (visible, linked both ways) ----
    data_ws = wb.create_sheet("Txn Detail")
    data_ws.sheet_view.showGridLines = False
    data_ws.tab_color = "9C0006" if n_flags else "1F3864"
    data_ws.merge_cells("A1:H1")
    data_ws["A1"] = "FULL POST-DISBURSEMENT TRANSACTION DETAIL"
    data_ws["A1"].font = TITLE_FONT
    data_ws["A1"].fill = TITLE_FILL
    data_ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    back_cell = data_ws["A2"]
    back_cell.value = "←  Back to Summary"
    back_cell.hyperlink = "#'EUF Summary'!A1"
    back_cell.font = Font(name=FONT_NAME, size=9, bold=True, color="0563C1", underline="single")

    post = df[df["Days_From_Disbursement"] >= 0][
        ["sDate", "sAmount", "sCreditOrDebit", "sNarration", "sCategory",
         "End_Use_Bucket", "Red_Flag", "Flag_Reason"]].copy()
    post.columns = ["Date", "Amount", "Cr/Dr", "Narration", "Bank Category",
                    "End-Use Bucket", "Red Flag", "Flag Reason"]
    header_row = 4
    for j, col in enumerate(post.columns):
        c = data_ws.cell(row=header_row, column=1 + j, value=col)
        c.font = SECTION_FONT
        c.fill = SECTION_FILL
        c.border = BORDER
    first_data_row = header_row + 1
    for i, (_, rrow) in enumerate(post.iterrows()):
        rr = first_data_row + i
        for j, col in enumerate(post.columns):
            val = rrow[col]
            c = data_ws.cell(row=rr, column=1 + j, value=(val.date() if col == "Date" else val))
            c.font = VALUE_FONT
            c.border = BORDER
            c.alignment = WRAP
            if col == "Date":
                c.number_format = "dd-mmm-yyyy"
            if col == "Amount":
                c.number_format = "#,##0.00"
        if i % 2 == 1:
            for j in range(len(post.columns)):
                data_ws.cell(row=rr, column=1 + j).fill = PatternFill("solid", fgColor=GREY)
    last_data_row = first_data_row + len(post) - 1
    if last_data_row >= first_data_row:
        data_ws.conditional_formatting.add(
            f"A{first_data_row}:H{last_data_row}",
            CellIsRule(operator="equal", formula=[f"$G{first_data_row}=TRUE"],
                       fill=PatternFill("solid", fgColor=RED_FILL))
        )
    data_widths = {"A": 14, "B": 14, "C": 10, "D": 42, "E": 20, "F": 26, "G": 10, "H": 45}
    for col, w in data_widths.items():
        data_ws.column_dimensions[col].width = w
    data_ws.freeze_panes = f"A{first_data_row}"
    data_ws.page_setup.orientation = "landscape"
    data_ws.page_setup.fitToWidth = 1
    data_ws.sheet_properties.pageSetUpPr.fitToPage = True

    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    p = argparse.ArgumentParser(description="Build a EUF summary directly from a full BSA report")
    p.add_argument("--input", required=True)
    p.add_argument("--loan-amount", type=float, default=None)
    p.add_argument("--loan-type", default="LAP", choices=["LAP", "HL"])
    p.add_argument("--lender-hint", default=None)
    p.add_argument("--loan-account-no", default="")
    p.add_argument("--purpose", default="")
    p.add_argument("--window-days", type=int, default=90)
    p.add_argument("--out", default="euf_from_bsa.xlsx")
    args = p.parse_args()

    build(args.input, args.loan_amount, args.loan_type, args.lender_hint,
          args.loan_account_no, args.purpose, args.out, args.window_days)
    print("Saved:", args.out)
