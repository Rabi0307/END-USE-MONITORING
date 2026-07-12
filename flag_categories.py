"""
flag_categories.py
--------------------
Flags every distinct sCategory value found in a category-list export
(a single-column file like S_Category.xlsx) against the EUF classification
rules in end_use_engine.py -- showing bucket, risk level, permitted status
for LAP/HL, and whether it's a bounce/dishonour event, plus how often each
category occurred in the file.
"""
import sys
from collections import Counter
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule

from end_use_engine import resolve_category, BOUNCE_CATEGORIES

FONT_NAME = "Arial"
NAVY, RED_FILL, AMBER_FILL, GREEN_FILL, GREY = "1F3864", "FFC7CE", "FFF2CC", "C6EFCE", "F2F2F2"
HEADER_FONT = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
NORMAL_FONT = Font(name=FONT_NAME, size=10)
TITLE_FONT = Font(name=FONT_NAME, size=14, bold=True, color="FFFFFF")
BORDER = Border(left=Side(style="thin", color="B7B7B7"), right=Side(style="thin", color="B7B7B7"),
                 top=Side(style="thin", color="B7B7B7"), bottom=Side(style="thin", color="B7B7B7"))


def flag_category_file(input_path, out_path, sheet_name=None, column_name="sCategory"):
    wb_in = openpyxl.load_workbook(input_path, data_only=True)
    ws_in = wb_in[sheet_name] if sheet_name else wb_in[wb_in.sheetnames[0]]
    rows = list(ws_in.iter_rows(values_only=True))
    header = rows[0]
    col_idx = header.index(column_name) if column_name in header else 0
    counts = Counter(r[col_idx] for r in rows[1:] if r and r[col_idx] is not None)

    results = []
    for cat, cnt in counts.items():
        info = resolve_category(cat)
        is_bounce = cat in BOUNCE_CATEGORIES or (isinstance(cat, str) and any(b in cat for b in BOUNCE_CATEGORIES))
        results.append({
            "Category": cat, "Occurrences": cnt, "End-Use Bucket": info["Bucket"],
            "Risk Level": info["Risk"], "Permitted (LAP)": info["LAP"], "Permitted (HL)": info["HL"],
            "Bounce/Dishonour Event": "YES" if is_bounce else "",
        })
    risk_rank = {"High": 0, "Medium": 1, "Low": 2}
    results.sort(key=lambda r: (risk_rank.get(r["Risk Level"], 1), -r["Occurrences"]))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Category Flags"
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:G1")
    ws["A1"] = "CATEGORY FLAGGING — END-USE CLASSIFICATION REVIEW"
    ws["A1"].font, ws["A1"].fill = TITLE_FONT, PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 24

    n_bounce = sum(r["Occurrences"] for r in results if r["Bounce/Dishonour Event"] == "YES")
    n_total = sum(r["Occurrences"] for r in results)
    n_unmapped = sum(1 for r in results if r["End-Use Bucket"] == "Uncategorised")
    ws["A3"] = f"Total transactions: {n_total:,}  |  Distinct categories: {len(results)}  |  Bounce/dishonour events: {n_bounce}  |  Unmapped categories: {n_unmapped}"
    ws["A3"].font = Font(name=FONT_NAME, size=10, italic=True, color="808080")
    ws.merge_cells("A3:G3")

    headers = ["Category", "Occurrences", "End-Use Bucket", "Risk Level", "Permitted (LAP)", "Permitted (HL)", "Bounce/Dishonour Event"]
    r0 = 5
    for j, h in enumerate(headers):
        c = ws.cell(row=r0, column=1 + j, value=h)
        c.font, c.fill, c.border = HEADER_FONT, PatternFill("solid", fgColor=NAVY), BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[r0].height = 26

    for i, row in enumerate(results):
        rr = r0 + 1 + i
        for j, h in enumerate(headers):
            c = ws.cell(row=rr, column=1 + j, value=row[h])
            c.font, c.border = NORMAL_FONT, BORDER
        if i % 2 == 1:
            for j in range(len(headers)):
                ws.cell(row=rr, column=1 + j).fill = PatternFill("solid", fgColor=GREY)

    last_row = r0 + len(results)
    risk_col = headers.index("Risk Level") + 1
    from openpyxl.utils import get_column_letter
    risk_letter = get_column_letter(risk_col)
    ws.conditional_formatting.add(f"{risk_letter}{r0+1}:{risk_letter}{last_row}",
        CellIsRule(operator="equal", formula=['"High"'], fill=PatternFill("solid", fgColor=RED_FILL)))
    ws.conditional_formatting.add(f"{risk_letter}{r0+1}:{risk_letter}{last_row}",
        CellIsRule(operator="equal", formula=['"Medium"'], fill=PatternFill("solid", fgColor=AMBER_FILL)))
    ws.conditional_formatting.add(f"{risk_letter}{r0+1}:{risk_letter}{last_row}",
        CellIsRule(operator="equal", formula=['"Low"'], fill=PatternFill("solid", fgColor=GREEN_FILL)))

    bounce_col = headers.index("Bounce/Dishonour Event") + 1
    bounce_letter = get_column_letter(bounce_col)
    ws.conditional_formatting.add(f"{bounce_letter}{r0+1}:{bounce_letter}{last_row}",
        CellIsRule(operator="equal", formula=['"YES"'], fill=PatternFill("solid", fgColor=RED_FILL)))

    widths = {"A": 34, "B": 13, "C": 34, "D": 11, "E": 15, "F": 13, "G": 20}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = f"A{r0+1}"

    wb.save(out_path)
    return out_path, results


if __name__ == "__main__":
    inp = sys.argv[1] if len(sys.argv) > 1 else "S_Category.xlsx"
    out = sys.argv[2] if len(sys.argv) > 2 else "Category_Flags.xlsx"
    out_path, results = flag_category_file(inp, out)
    print(f"Saved: {out_path}  ({len(results)} categories flagged)")
