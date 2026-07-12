"""
build_onepager.py
------------------
A condensed, ONE-PAGE version of the EUF monitoring format for a single loan
-- everything that matters (loan details, utilisation vs declared purpose,
red flags, sign-off) on a single printable A4 landscape page, PLUS the full
transaction detail one click away on its own visible sheet ("Txn Detail")
in the same workbook -- linked both ways so you can jump to it and back.

Use this when the full 10-sheet workbook is more than you need per file
(e.g. reviewing hundreds of loans quickly) -- pair it with batch_report.py
for the portfolio-level view, and fall back to build_excel.py's full
workbook only for loans that need deeper investigation.

USAGE
    python3 build_onepager.py --input statement.xlsx --loan-amount 892242 \
        --loan-type LAP --lender-hint SBFC --borrower "Name" --out onepager.xlsx
"""
import argparse
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule

from end_use_engine import run as run_engine

FONT_NAME = "Arial"
NAVY = "1F3864"
LIGHT_BLUE = "DCE6F1"
YELLOW = "FFF2CC"
RED_FILL = "FFC7CE"
GREEN_FILL = "C6EFCE"
GREY = "F2F2F2"

TITLE_FONT = Font(name=FONT_NAME, size=13, bold=True, color="FFFFFF")
SECTION_FONT = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
LABEL_FONT = Font(name=FONT_NAME, size=9, bold=True, color=NAVY)
VALUE_FONT = Font(name=FONT_NAME, size=9)
NOTE_FONT = Font(name=FONT_NAME, size=8, italic=True, color="808080")
TITLE_FILL = PatternFill("solid", fgColor=NAVY)
SECTION_FILL = PatternFill("solid", fgColor=NAVY)
INPUT_FILL = PatternFill("solid", fgColor=YELLOW)
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")


def section_header(ws, cell_range, text):
    ws.merge_cells(cell_range)
    c = ws[cell_range.split(":")[0]]
    c.value = text
    c.font = SECTION_FONT
    c.fill = SECTION_FILL
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[c.row].height = 16


def label_value(ws, label_cell, value_cell, label, value, number_format=None):
    ws[label_cell] = label
    ws[label_cell].font = LABEL_FONT
    ws[value_cell] = value
    ws[value_cell].font = VALUE_FONT
    ws[value_cell].fill = INPUT_FILL if value in ("", None) else PatternFill(fill_type=None)
    if number_format:
        ws[value_cell].number_format = number_format
    for cell in (label_cell, value_cell):
        ws[cell].border = BORDER


def build_onepager(statement_path, loan_amount, loan_type, lender_hint,
                    borrower_name, loan_account_no, purpose_declared,
                    property_address, out_path, window_days=90):

    df, disb_row, summary = run_engine(statement_path, loan_amount=loan_amount,
                                        loan_type=loan_type, lender_hint=lender_hint,
                                        window_days=window_days)
    disb_amt = float(disb_row["sAmount"]) if disb_row is not None else loan_amount
    n_flags = int(df["Red_Flag"].sum())
    flagged_amt = float(df.loc[df["Red_Flag"], "sAmount"].sum())
    top_flags = df[df["Red_Flag"]].sort_values("sAmount", ascending=False).head(5)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EUF One-Pager"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    ws["A1"] = "POST-DISBURSEMENT END-USE OF FUNDS — ONE-PAGE SUMMARY"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 22

    # ---- LOAN DETAILS BLOCK ----
    section_header(ws, "A3:H3", "LOAN DETAILS")
    label_value(ws, "A4", "B4", "Borrower", borrower_name)
    label_value(ws, "C4", "D4", "Loan A/C No.", loan_account_no)
    label_value(ws, "E4", "F4", "Loan Type", loan_type)
    label_value(ws, "G4", "H4", "Sanctioned/Disbursed (Rs.)", disb_amt, "#,##0")
    label_value(ws, "A5", "B5", "Disbursement Date", disb_row["sDate"].date() if disb_row is not None else None, "dd-mmm-yyyy")
    label_value(ws, "C5", "D5", "Lender / NBFC", lender_hint or "")
    label_value(ws, "E5", "F5", "Monitoring Window", f"{window_days} days")
    label_value(ws, "G5", "H5", "Window Ends", summary["window_end"].date(), "dd-mmm-yyyy")
    ws.merge_cells("B6:H6")
    label_value(ws, "A6", "B6", "Declared Purpose", purpose_declared)

    # ---- UTILISATION SUMMARY (left) + RED FLAG SUMMARY (right) ----
    section_header(ws, "A8:D8", "UTILISATION SUMMARY (vs. Disbursed Amount)")
    section_header(ws, "E8:H8", "RED-FLAG SUMMARY")

    headers = ["End-Use Bucket", "Amount (Rs.)", "%", "Assessment"]
    for j, h in enumerate(headers):
        c = ws.cell(row=9, column=1 + j, value=h)
        c.font = LABEL_FONT
        c.border = BORDER
        c.fill = PatternFill("solid", fgColor=LIGHT_BLUE)

    bucket_df = summary["by_bucket"].head(6)
    r = 10
    for _, row in bucket_df.iterrows():
        pct = row["Pct_of_Loan_Amount"]
        assessment = "Review" if pct > 0.10 else "OK"
        ws.cell(row=r, column=1, value=row["End_Use_Bucket"]).font = VALUE_FONT
        ws.cell(row=r, column=2, value=row["Total_Amount"]).number_format = "#,##0"
        ws.cell(row=r, column=3, value=pct).number_format = "0.0%"
        acell = ws.cell(row=r, column=4, value=assessment)
        for c in range(1, 5):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).font = VALUE_FONT
        r += 1
    last_bucket_row = r - 1

    ws.conditional_formatting.add(f"D10:D{last_bucket_row}",
        CellIsRule(operator="equal", formula=['"OK"'], fill=PatternFill("solid", fgColor=GREEN_FILL)))
    ws.conditional_formatting.add(f"D10:D{last_bucket_row}",
        CellIsRule(operator="equal", formula=['"Review"'], fill=PatternFill("solid", fgColor=RED_FILL)))

    # Red flag summary (right side, rows 9-15)
    ws["E9"] = "Total Red-Flagged Txns"
    ws["E9"].font = LABEL_FONT
    ws["F9"] = n_flags
    ws["F9"].font = Font(name=FONT_NAME, size=11, bold=True,
                          color="9C0006" if n_flags else "006100")
    ws["G9"] = "Flagged Amount (Rs.)"
    ws["G9"].font = LABEL_FONT
    ws["H9"] = flagged_amt
    ws["H9"].number_format = "#,##0"
    ws["H9"].font = Font(name=FONT_NAME, size=10, bold=True)

    ws.merge_cells("E10:H10")
    ws["E10"] = "Top flagged transactions:"
    ws["E10"].font = LABEL_FONT
    rr = 11
    if top_flags.empty:
        ws.merge_cells(f"E{rr}:H{rr}")
        ws[f"E{rr}"] = "None — no automated red flags in this window."
        ws[f"E{rr}"].font = VALUE_FONT
    else:
        for _, fr in top_flags.iterrows():
            ws.merge_cells(f"E{rr}:H{rr}")
            txt = f"{fr['sDate'].date()} — Rs.{fr['sAmount']:,.0f} — {fr['Flag_Reason']}"
            c = ws[f"E{rr}"]
            c.value = txt
            c.font = Font(name=FONT_NAME, size=8)
            c.fill = PatternFill("solid", fgColor=RED_FILL)
            c.alignment = WRAP
            rr += 1

    status_row = max(last_bucket_row, rr) + 1
    ws.merge_cells(f"A{status_row}:H{status_row}")
    overall = "REVIEW REQUIRED" if n_flags > 0 else "CLEAN"
    ws[f"A{status_row}"] = f"OVERALL STATUS: {overall}"
    ws[f"A{status_row}"].font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
    ws[f"A{status_row}"].fill = PatternFill("solid", fgColor="9C0006" if overall == "REVIEW REQUIRED" else "006100")
    ws[f"A{status_row}"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[status_row].height = 18

    # ---- SIGN-OFF ----
    signoff_row = status_row + 2
    section_header(ws, f"A{signoff_row}:H{signoff_row}", "FINDING & SIGN-OFF")
    r = signoff_row + 1
    ws.merge_cells(f"A{r}:H{r}")
    ws[f"A{r}"] = "Finding (strike out as applicable):  Satisfactory  /  Requires Further Clarification  /  Diversion Suspected"
    ws[f"A{r}"].font = VALUE_FONT
    r += 1
    ws.merge_cells(f"A{r}:D{r+1}")
    ws[f"A{r}"] = "Remarks:"
    ws[f"A{r}"].font = LABEL_FONT
    ws[f"A{r}"].fill = INPUT_FILL
    ws[f"A{r}"].alignment = WRAP
    ws.merge_cells(f"E{r}:H{r+1}")
    ws[f"E{r}"].fill = INPUT_FILL
    r += 3
    ws[f"A{r}"] = "Verified By / Date:"
    ws[f"A{r}"].font = LABEL_FONT
    ws.merge_cells(f"B{r}:D{r}")
    ws[f"B{r}"].border = Border(bottom=THIN)
    ws[f"E{r}"] = "Approved By / Date:"
    ws[f"E{r}"].font = LABEL_FONT
    ws.merge_cells(f"F{r}:H{r}")
    ws[f"F{r}"].border = Border(bottom=THIN)

    # ---- link to transaction detail, from the summary page ----
    link_row = r + 2
    ws.merge_cells(f"A{link_row}:H{link_row}")
    link_cell = ws[f"A{link_row}"]
    link_cell.value = "View full transaction detail  →  see the 'Txn Detail' tab (bottom of workbook)"
    link_cell.hyperlink = "#'Txn Detail'!A1"
    link_cell.font = Font(name=FONT_NAME, size=9, bold=True, color="0563C1", underline="single")
    link_cell.alignment = Alignment(horizontal="left")

    # column widths for a clean single printed page
    widths = {"A": 20, "B": 16, "C": 16, "D": 14, "E": 20, "F": 14, "G": 18, "H": 16}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.page_margins.top = ws.page_margins.bottom = 0.4

    # ---- TRANSACTION DETAIL SHEET (visible, one click away — not hidden) ----
    data_ws = wb.create_sheet("Txn Detail")
    data_ws.sheet_view.showGridLines = False
    data_ws.tab_color = "9C0006" if n_flags else "1F3864"

    data_ws.merge_cells("A1:H1")
    data_ws["A1"] = "FULL POST-DISBURSEMENT TRANSACTION DETAIL"
    data_ws["A1"].font = TITLE_FONT
    data_ws["A1"].fill = TITLE_FILL
    data_ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    data_ws.row_dimensions[1].height = 20

    back_cell = data_ws["A2"]
    back_cell.value = "←  Back to One-Page Summary"
    back_cell.hyperlink = "#'EUF One-Pager'!A1"
    back_cell.font = Font(name=FONT_NAME, size=9, bold=True, color="0563C1", underline="single")

    post = df[df["Days_From_Disbursement"] >= 0][
        ["sDate", "sAmount", "sCreditOrDebit", "sNarration", "sCategory",
         "End_Use_Bucket", "Red_Flag", "Flag_Reason"]].copy()
    post.columns = ["Date", "Amount", "Cr/Dr", "Narration", "Bank Category",
                    "End-Use Bucket", "Red Flag", "Flag Reason"]

    header_row = 4
    for j, col in enumerate(post.columns):
        c = data_ws.cell(row=header_row, column=1 + j, value=col)
        c.font = SECTION_FONT
        c.fill = SECTION_FILL
        c.border = BORDER
    first_data_row = header_row + 1
    for i, (_, rrow) in enumerate(post.iterrows()):
        rr = first_data_row + i
        for j, col in enumerate(post.columns):
            val = rrow[col]
            c = data_ws.cell(row=rr, column=1 + j, value=(val.date() if col == "Date" else val))
            c.font = VALUE_FONT
            c.border = BORDER
            c.alignment = WRAP
            if col == "Date":
                c.number_format = "dd-mmm-yyyy"
            if col == "Amount":
                c.number_format = "#,##0.00"
        if i % 2 == 1:
            for j in range(len(post.columns)):
                data_ws.cell(row=rr, column=1 + j).fill = PatternFill("solid", fgColor=GREY)
    last_data_row = first_data_row + len(post) - 1
    if last_data_row >= first_data_row:
        data_ws.conditional_formatting.add(
            f"A{first_data_row}:H{last_data_row}",
            CellIsRule(operator="equal", formula=[f"$G{first_data_row}=TRUE"],
                       fill=PatternFill("solid", fgColor=RED_FILL))
        )
        # Note: openpyxl anchors the formula's row references to the first cell of the
        # range and Excel auto-adjusts them per-row when applied, matching each row's own
        # "Red Flag" column (G) — same mechanism used in the full workbook's Txn Register.

    data_widths = {"A": 14, "B": 14, "C": 10, "D": 42, "E": 20, "F": 26, "G": 10, "H": 45}
    for col, w in data_widths.items():
        data_ws.column_dimensions[col].width = w
    data_ws.freeze_panes = f"A{first_data_row}"
    data_ws.page_setup.orientation = "landscape"
    data_ws.page_setup.fitToWidth = 1
    data_ws.page_setup.fitToHeight = 0
    data_ws.sheet_properties.pageSetUpPr.fitToPage = True

    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    p = argparse.ArgumentParser(description="Build a one-page EUF summary for a single loan")
    p.add_argument("--input", required=True)
    p.add_argument("--loan-amount", type=float, default=None)
    p.add_argument("--loan-type", default="LAP", choices=["LAP", "HL"])
    p.add_argument("--lender-hint", default=None)
    p.add_argument("--borrower", default="")
    p.add_argument("--loan-account-no", default="")
    p.add_argument("--purpose", default="")
    p.add_argument("--property-address", default="")
    p.add_argument("--window-days", type=int, default=90)
    p.add_argument("--out", default="onepager.xlsx")
    args = p.parse_args()

    build_onepager(args.input, args.loan_amount, args.loan_type, args.lender_hint,
                   args.borrower, args.loan_account_no, args.purpose,
                   args.property_address, args.out, args.window_days)
    print("Saved:", args.out)
