"""
run_euf_scan.py
-----------------
Standalone, single-file version — no other files needed.

WHAT TO DO
1. Copy this one file into the folder that contains your statements
   (either directly, e.g. folder/loan1.xlsx, or one-subfolder-per-loan,
   e.g. folder/SBFCLAP0000376039/report.xlsx — both work).
2. Open a terminal / command prompt IN that folder.
3. Run:
       pip install openpyxl pandas --break-system-packages
       python run_euf_scan.py

That's it — no arguments needed. It scans the folder it's sitting in,
uses each subfolder's name as the Loan Account Number (LAN), and names
the output report after the folder itself
(e.g. running it inside "reports_download" produces "reports_download_EUF_Report.xlsx").

Optional: put a "loan_master.csv" file in the same folder (columns:
filename, loan_amount, loan_type, lender_hint, borrower_name,
loan_account_no) for accurate loan amounts instead of auto-detection.
`filename` can match either the statement's file name or its parent
folder name (the LAN).
"""

import os
import sys
import csv
from datetime import timedelta, datetime

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

# ===========================================================================
# 1. END-USE CLASSIFICATION RULES  (edit this to match your own policy)
# ===========================================================================
END_USE_MAP = {
    "Direct Deposit":            {"Bucket": "Incoming Funds / Not Applicable", "Risk": "Low"},
    "Cheque Deposit":            {"Bucket": "Incoming Funds / Not Applicable", "Risk": "Low"},
    "Cash Deposit":               {"Bucket": "Incoming Funds / Not Applicable", "Risk": "Low"},
    "Refunds":                   {"Bucket": "Incoming Funds / Not Applicable", "Risk": "Low"},
    "Interest Received":         {"Bucket": "Incoming Funds / Not Applicable", "Risk": "Low"},
    "Dividends":                  {"Bucket": "Incoming Funds / Not Applicable", "Risk": "Low"},
    "Investment Income":          {"Bucket": "Incoming Funds / Not Applicable", "Risk": "Low"},
    "Rent Received":              {"Bucket": "Incoming Funds / Not Applicable", "Risk": "Low"},
    "Salary Received":            {"Bucket": "Incoming Funds / Not Applicable", "Risk": "Low"},
    "Subsidy":                    {"Bucket": "Incoming Funds / Not Applicable", "Risk": "Low"},
    "Loan Disbursed":             {"Bucket": "Other Loan Disbursement (New Debt)", "Risk": "High"},
    "EMI and Loans":             {"Bucket": "Debt Servicing (Existing Loans)", "Risk": "Medium"},
    "Emi_Paid":                  {"Bucket": "Debt Servicing (Existing Loans)", "Risk": "Medium"},
    "Home Loan":                   {"Bucket": "Debt Servicing (Existing Loans)", "Risk": "Medium"},
    "Card Payment":                {"Bucket": "Debt Servicing (Existing Loans)", "Risk": "Medium"},
    "Bounced I/W ECS":             {"Bucket": "Bounced Payment / Dishonour", "Risk": "High"},
    "Bounced I/W ECS Charges":     {"Bucket": "Bounced Payment / Dishonour", "Risk": "High"},
    "Bounced O/W Cheque":          {"Bucket": "Bounced Payment / Dishonour", "Risk": "High"},
    "Cash Withdrawals":          {"Bucket": "Cash Withdrawal",                 "Risk": "High"},
    "Cheque Withdrawal":           {"Bucket": "Cash Withdrawal",                 "Risk": "High"},
    "Gold":                      {"Bucket": "Investment / Gold Purchase",      "Risk": "High"},
    "Investment Expense":          {"Bucket": "Investment / Gold Purchase",      "Risk": "High"},
    "Fixed Deposit":               {"Bucket": "Investment / Gold Purchase",      "Risk": "Medium"},
    "Groceries and Shopping":    {"Bucket": "Household / Personal Spend",      "Risk": "Medium"},
    "Shopping":                   {"Bucket": "Household / Personal Spend",      "Risk": "Medium"},
    "Utilities and Bills":       {"Bucket": "Household / Personal Spend",      "Risk": "Medium"},
    "Dining Restaurant and Entertainment": {"Bucket": "Household / Personal Spend", "Risk": "Medium"},
    "Entertainment":               {"Bucket": "Household / Personal Spend",      "Risk": "Medium"},
    "Hotel and Travel":            {"Bucket": "Household / Personal Spend",      "Risk": "Medium"},
    "Medical":                     {"Bucket": "Household / Personal Spend",      "Risk": "Low"},
    "Insurance Premium":           {"Bucket": "Household / Personal Spend",      "Risk": "Low"},
    "Rent Paid":                   {"Bucket": "Household / Personal Spend",      "Risk": "Medium"},
    "Fuel":                      {"Bucket": "Household / Personal Spend",      "Risk": "Low"},
    "Bank Fees and Charges":     {"Bucket": "Bank Charges",                    "Risk": "Low"},
    "Bank_fees_and_Charges":     {"Bucket": "Bank Charges",                    "Risk": "Low"},
    "External Transfers":        {"Bucket": "Third-Party Transfer (Unverified)", "Risk": "Medium"},
    "Transfer to UPI":           {"Bucket": "Third-Party Transfer (Unverified)", "Risk": "Medium"},
}
DEFAULT_BUCKET = {"Bucket": "Uncategorised", "Risk": "Medium"}
BOUNCE_CATEGORIES = {"Bounced I/W ECS", "Bounced I/W ECS Charges", "Bounced O/W Cheque"}


def resolve_category(cat):
    """Look up END_USE_MAP, handling multi-label values (e.g. 'A,B') and
    falling back to DEFAULT_BUCKET (never silently 'safe') if unmapped."""
    if cat in END_USE_MAP:
        return END_USE_MAP[cat]
    if isinstance(cat, str) and "," in cat:
        for part in cat.split(","):
            part = part.strip()
            if part in END_USE_MAP:
                return END_USE_MAP[part]
    return DEFAULT_BUCKET



CASH_WITHDRAWAL_FLAG_PCT  = 0.05   # single cash withdrawal > 5% of loan amount
LARGE_TXN_FLAG_PCT        = 0.10   # any single debit > 10% of loan amount
SAME_DAY_OUTFLOW_FLAG_PCT = 0.15   # outflow within 24h of disbursement > 15% of loan
MIN_FLAG_AMOUNT           = 1000   # ignore trivially small txns (e.g. gold SIPs) when flagging
WINDOW_DAYS               = 90     # monitoring window post-disbursement
DEFAULT_LOAN_TYPE         = "LAP"  # used when a file isn't in loan_master.csv


# ===========================================================================
# 2. LOAD + CLASSIFY ONE STATEMENT
# ===========================================================================
def load_statement(path):
    xl = pd.ExcelFile(path)
    sheet_names_lower = {s.lower(): s for s in xl.sheet_names}
    df = pd.read_excel(path, sheet_name=0)
    for candidate in ("allaccountxns", "accountxns"):
        if candidate in sheet_names_lower:
            df = pd.read_excel(path, sheet_name=sheet_names_lower[candidate])
            break
    else:
        for sheet in xl.sheet_names:
            if "bank account" in sheet.lower() or "transaction" in sheet.lower():
                df = pd.read_excel(path, sheet_name=sheet)
                break
    df.columns = [c.strip() for c in df.columns]
    if "is_excluded" in df.columns:
        df = df[df["is_excluded"] != True]  # noqa: E712
    df["sDate"] = pd.to_datetime(df["sDate"])
    df["sAmount"] = pd.to_numeric(df["sAmount"], errors="coerce")
    return df.sort_values("sDate").reset_index(drop=True)


def detect_disbursement(df, loan_amount=None, lender_hint=None, tolerance=0.02):
    credits = df[df["sCreditOrDebit"].str.upper() == "CREDIT"].copy()
    if lender_hint:
        mask = credits["sNarration"].str.contains(lender_hint, case=False, na=False)
        if mask.any():
            credits = credits[mask]
    if loan_amount:
        credits["_diff"] = (credits["sAmount"] - loan_amount).abs() / loan_amount
        candidate = credits[credits["_diff"] <= tolerance].sort_values("_diff")
        if not candidate.empty:
            return candidate.iloc[0]
    if credits.empty:
        return None
    return credits.sort_values("sAmount", ascending=False).iloc[0]


def analyze_file(path, loan_amount=None, lender_hint=None, window_days=WINDOW_DAYS):
    df = load_statement(path)
    disb_row = detect_disbursement(df, loan_amount, lender_hint)
    if disb_row is None:
        raise ValueError("No CREDIT transactions found — cannot identify a disbursement.")
    effective_amount = float(loan_amount) if loan_amount else float(disb_row["sAmount"])
    disb_date = disb_row["sDate"]
    window_end = disb_date + timedelta(days=window_days)

    df["End_Use_Bucket"] = df["sCategory"].apply(lambda c: resolve_category(c)["Bucket"])
    df["Days_From_Disbursement"] = (df["sDate"] - disb_date).dt.days

    flags, reasons = [], []
    for _, r in df.iterrows():
        reason = []
        cat = r["sCategory"]
        is_bounce = cat in BOUNCE_CATEGORIES or (isinstance(cat, str) and any(b in cat for b in BOUNCE_CATEGORIES))
        if is_bounce:
            reason.append(f"Cheque/ECS bounce or dishonour event ('{cat}') — review immediately")
        if r["sCreditOrDebit"].upper() == "DEBIT" and r["Days_From_Disbursement"] >= 0 and effective_amount:
            pct = r["sAmount"] / effective_amount
            if cat == "Cash Withdrawals" and pct >= CASH_WITHDRAWAL_FLAG_PCT:
                reason.append(f"Cash withdrawal = {pct:.1%} of loan amount")
            if pct >= LARGE_TXN_FLAG_PCT:
                reason.append(f"Single debit = {pct:.1%} of loan amount (>{LARGE_TXN_FLAG_PCT:.0%} threshold)")
            if 0 <= r["Days_From_Disbursement"] <= 1 and pct >= SAME_DAY_OUTFLOW_FLAG_PCT:
                reason.append("Large outflow within 24 hrs of disbursement")
            if cat == "Gold" and r["sAmount"] >= MIN_FLAG_AMOUNT:
                reason.append("Funds routed to gold/investment purchase")
        flags.append(bool(reason))
        reasons.append("; ".join(reason))
    df["Red_Flag"] = flags
    df["Flag_Reason"] = reasons

    post = df[(df["Days_From_Disbursement"] >= 0) & (df["sDate"] <= window_end) &
              (df["sCreditOrDebit"].str.upper() == "DEBIT")]
    by_bucket = (post.groupby("End_Use_Bucket")["sAmount"].agg(["sum", "count"])
                 .reset_index().rename(columns={"sum": "Total_Amount", "count": "Txn_Count"})
                 .sort_values("Total_Amount", ascending=False))
    by_bucket["Pct_of_Loan_Amount"] = by_bucket["Total_Amount"] / effective_amount if effective_amount else 0

    return {
        "df": df, "disb_row": disb_row, "effective_amount": effective_amount,
        "window_end": window_end, "by_bucket": by_bucket, "post": post,
    }


# ===========================================================================
# 3. FIND STATEMENT FILES + LOAN MASTER MAPPING (folder name = LAN)
# ===========================================================================
def find_statement_files(base_folder, this_script_name):
    results = []
    for root, dirs, files in os.walk(base_folder):
        for f in files:
            if f.lower().endswith((".xlsx", ".xls")) and not f.startswith("~$") and f != this_script_name:
                full = os.path.join(root, f)
                rel = os.path.relpath(full, base_folder)
                parent = os.path.basename(root)
                folder_name = parent if os.path.abspath(root) != os.path.abspath(base_folder) else ""
                results.append({"path": full, "rel_path": rel, "folder_name": folder_name, "filename": f})
    results.sort(key=lambda r: r["rel_path"])
    return results


def load_mapping(base_folder):
    mapping_path = os.path.join(base_folder, "loan_master.csv")
    if not os.path.exists(mapping_path):
        return {}
    mapping = {}
    with open(mapping_path, newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            key = (row.get("filename") or "").strip()
            if key:
                mapping[key] = row
    print(f"Loaded loan_master.csv with {len(mapping)} row(s)")
    return mapping


# ===========================================================================
# 4. EXCEL REPORT WRITER
# ===========================================================================
FONT_NAME = "Arial"
NAVY, RED_FILL, GREEN_FILL, GREY = "1F3864", "FFC7CE", "C6EFCE", "F2F2F2"
HEADER_FONT = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
NORMAL_FONT = Font(name=FONT_NAME, size=10)
TITLE_FONT = Font(name=FONT_NAME, size=14, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
STRIPE_FILL = PatternFill("solid", fgColor=GREY)
BORDER = Border(*(Side(style="thin", color="B7B7B7"),) * 4) if False else Border(
    left=Side(style="thin", color="B7B7B7"), right=Side(style="thin", color="B7B7B7"),
    top=Side(style="thin", color="B7B7B7"), bottom=Side(style="thin", color="B7B7B7"))
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_title(ws, cell_range, text):
    ws.merge_cells(cell_range)
    c = ws[cell_range.split(":")[0]]
    c.value, c.font, c.fill = text, TITLE_FONT, TITLE_FILL
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[c.row].height = 26


def style_header_row(ws, row, ncols):
    for i in range(1, ncols + 1):
        c = ws.cell(row=row, column=i)
        c.font, c.fill, c.alignment, c.border = HEADER_FONT, HEADER_FILL, CENTER, BORDER
    ws.row_dimensions[row].height = 30


def write_table(ws, headers, rows, start_row, currency_cols=None, pct_cols=None, date_cols=None):
    currency_cols, pct_cols, date_cols = currency_cols or [], pct_cols or [], date_cols or []
    for j, h in enumerate(headers):
        ws.cell(row=start_row, column=1 + j, value=h)
    style_header_row(ws, start_row, len(headers))
    for i, row in enumerate(rows):
        rr = start_row + 1 + i
        for j, h in enumerate(headers):
            c = ws.cell(row=rr, column=1 + j, value=row.get(h))
            c.font, c.border, c.alignment = NORMAL_FONT, BORDER, WRAP
            if h in currency_cols: c.number_format = "#,##0.00"
            if h in pct_cols: c.number_format = "0.0%"
            if h in date_cols: c.number_format = "dd-mmm-yyyy"
        if i % 2 == 1:
            for j in range(len(headers)):
                ws.cell(row=rr, column=1 + j).fill = STRIPE_FILL
    return start_row + 1 + len(rows)


def autofit(ws, min_w=10, max_w=50):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                widths[cell.column_letter] = max(widths.get(cell.column_letter, 0), len(str(cell.value)))
    for col, w in widths.items():
        ws.column_dimensions[col].width = max(min_w, min(max_w, w + 2))


def write_report(results, all_flags, bucket_totals, errors, out_path, n_files):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # --- Dashboard ---
    ws = wb.create_sheet("Dashboard")
    style_title(ws, "A1:D1", "PORTFOLIO END-USE OF FUNDS — SUMMARY DASHBOARD")
    ws.sheet_view.showGridLines = False
    n_ok, n_flagged = len(results), sum(1 for r in results if r["Status"] == "REVIEW REQUIRED")
    total_disb = sum(r["Disbursed / Sanctioned Amount"] for r in results)
    total_flag_amt = sum(r["Red Flag Amount"] for r in results)
    kpis = [("Files Scanned", n_files), ("Loans Processed Successfully", n_ok),
            ("Files with Errors", len(errors)), ("Loans Flagged — Review Required", n_flagged),
            ("Loans Clean", n_ok - n_flagged), ("Total Disbursed / Sanctioned (Rs.)", total_disb),
            ("Total Red-Flagged Amount (Rs.)", total_flag_amt),
            ("Report Generated", datetime.today().strftime("%d-%b-%Y %H:%M"))]
    r = 3
    for label, val in kpis:
        ws.cell(row=r, column=1, value=label).font = Font(name=FONT_NAME, bold=True, color=NAVY)
        ws.cell(row=r, column=1).border = BORDER
        c = ws.cell(row=r, column=2, value=val)
        c.border, c.font = BORDER, Font(name=FONT_NAME, size=11, bold=True)
        if "Rs." in label: c.number_format = "#,##0.00"
        r += 1
    ws.column_dimensions["A"].width, ws.column_dimensions["B"].width = 40, 24

    chart_row = r + 2
    ws.cell(row=chart_row, column=1, value="Portfolio Utilisation by End-Use Bucket").font = Font(bold=True, color=NAVY)
    chart_row += 1
    sorted_buckets = sorted(bucket_totals.items(), key=lambda x: -x[1])
    for i, (bucket, amt) in enumerate(sorted_buckets):
        ws.cell(row=chart_row + i, column=1, value=bucket)
        ws.cell(row=chart_row + i, column=2, value=amt).number_format = "#,##0"
    if sorted_buckets:
        chart = BarChart()
        chart.title, chart.type, chart.style = "Portfolio-Wide Utilisation by End-Use Bucket (Rs.)", "bar", 10
        data = Reference(ws, min_col=2, min_row=chart_row, max_row=chart_row + len(sorted_buckets) - 1)
        cats = Reference(ws, min_col=1, min_row=chart_row, max_row=chart_row + len(sorted_buckets) - 1)
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(cats)
        chart.legend, chart.height, chart.width = None, 9, 20
        ws.add_chart(chart, "D3")

    # --- Portfolio Summary ---
    ws = wb.create_sheet("Portfolio Summary")
    style_title(ws, "A1:N1", "PORTFOLIO SUMMARY — ONE ROW PER LOAN")
    ws.sheet_view.showGridLines = False
    headers = ["File", "LAN (folder name)", "Borrower", "Loan Type", "Amount Source", "Disbursement Date",
               "Disbursed / Sanctioned Amount", "Total Utilised (window)", "% Utilised",
               "Top End-Use Bucket", "Top Bucket %", "Red Flag Count", "Red Flag Amount", "Status"]
    last_row = write_table(ws, headers, results, 3,
                            currency_cols=["Disbursed / Sanctioned Amount", "Total Utilised (window)", "Red Flag Amount"],
                            pct_cols=["% Utilised", "Top Bucket %"], date_cols=["Disbursement Date"])
    status_col = get_column_letter(headers.index("Status") + 1)
    ws.conditional_formatting.add(f"{status_col}4:{status_col}{last_row-1}",
        CellIsRule(operator="equal", formula=['"CLEAN"'], fill=PatternFill("solid", fgColor=GREEN_FILL)))
    ws.conditional_formatting.add(f"{status_col}4:{status_col}{last_row-1}",
        CellIsRule(operator="equal", formula=['"REVIEW REQUIRED"'], fill=PatternFill("solid", fgColor=RED_FILL)))
    autofit(ws)
    ws.freeze_panes = "A4"

    # --- All Red Flags ---
    ws = wb.create_sheet("All Red Flags")
    style_title(ws, "A1:H1", "CONSOLIDATED RED-FLAG REGISTER — ACROSS ALL FILES")
    ws.sheet_view.showGridLines = False
    if all_flags:
        write_table(ws, ["File", "LAN (folder name)", "Borrower", "Date", "Amount", "Narration", "Counterparty", "Flag Reason"],
                    all_flags, 3, currency_cols=["Amount"], date_cols=["Date"])
    else:
        ws["A3"] = "No red-flagged transactions found across the portfolio."
        ws["A3"].font = NORMAL_FONT
    autofit(ws)
    ws.column_dimensions["F"].width = 45
    ws.column_dimensions["H"].width = 45
    ws.freeze_panes = "A4"

    # --- Category Aggregate ---
    ws = wb.create_sheet("Category Aggregate")
    style_title(ws, "A1:C1", "PORTFOLIO-WIDE UTILISATION BY END-USE BUCKET")
    ws.sheet_view.showGridLines = False
    total_all = sum(bucket_totals.values()) or 1
    rows = [{"End-Use Bucket": b, "Total Amount (Rs.)": a, "% of Portfolio Debits": a / total_all}
            for b, a in sorted_buckets]
    write_table(ws, ["End-Use Bucket", "Total Amount (Rs.)", "% of Portfolio Debits"], rows, 3,
                currency_cols=["Total Amount (Rs.)"], pct_cols=["% of Portfolio Debits"])
    autofit(ws)

    # --- Errors ---
    ws = wb.create_sheet("Errors")
    style_title(ws, "A1:B1", "FILES THAT COULD NOT BE PROCESSED")
    ws.sheet_view.showGridLines = False
    if errors:
        write_table(ws, ["File", "Error"], errors, 3)
    else:
        ws["A3"] = "No errors — all files processed successfully."
        ws["A3"].font = NORMAL_FONT
    autofit(ws)
    ws.column_dimensions["B"].width = 60

    order = ["Dashboard", "Portfolio Summary", "All Red Flags", "Category Aggregate", "Errors"]
    wb._sheets = [wb[n] for n in order]
    wb.active = 0
    for n in order:
        wb[n].page_setup.orientation = "landscape"
        wb[n].page_setup.fitToWidth = 1
        wb[n].page_setup.fitToHeight = 0
        wb[n].sheet_properties.pageSetUpPr.fitToPage = True

    wb.save(out_path)


# ===========================================================================
# 5. MAIN — scans the folder this script is run from
# ===========================================================================
def main():
    base_folder = os.getcwd()
    this_script_name = os.path.basename(__file__)
    folder_label = os.path.basename(os.path.abspath(base_folder)) or "portfolio"

    mapping = load_mapping(base_folder)
    files = find_statement_files(base_folder, this_script_name)
    print(f"Found {len(files)} Excel file(s) under: {base_folder}")

    results, all_flags, bucket_totals, errors = [], [], {}, []

    for i, entry in enumerate(files, 1):
        row_map = mapping.get(entry["filename"]) or mapping.get(entry["folder_name"]) or {}
        loan_amount = float(row_map["loan_amount"]) if row_map.get("loan_amount") else None
        loan_type = (row_map.get("loan_type") or DEFAULT_LOAN_TYPE).upper()
        lender_hint = row_map.get("lender_hint") or None
        borrower_name = row_map.get("borrower_name") or ""
        # LAN = the mapping's value if given, else the file's own subfolder name
        lan = row_map.get("loan_account_no") or entry["folder_name"] or ""

        try:
            res = analyze_file(entry["path"], loan_amount=loan_amount, lender_hint=lender_hint)
            df, disb_row = res["df"], res["disb_row"]
            effective_amount = res["effective_amount"]
            n_flags = int(df["Red_Flag"].sum())
            flagged_amount = float(df.loc[df["Red_Flag"], "sAmount"].sum())
            top_bucket_row = res["by_bucket"].iloc[0] if not res["by_bucket"].empty else None

            results.append({
                "File": entry["rel_path"], "LAN (folder name)": lan, "Borrower": borrower_name,
                "Loan Type": loan_type, "Amount Source": "Matched (mapping)" if row_map else "Auto-detected",
                "Disbursement Date": disb_row["sDate"].date(), "Disbursed / Sanctioned Amount": effective_amount,
                "Total Utilised (window)": res["post"]["sAmount"].sum(),
                "% Utilised": (res["post"]["sAmount"].sum() / effective_amount) if effective_amount else 0,
                "Top End-Use Bucket": top_bucket_row["End_Use_Bucket"] if top_bucket_row is not None else "",
                "Top Bucket %": top_bucket_row["Pct_of_Loan_Amount"] if top_bucket_row is not None else None,
                "Red Flag Count": n_flags, "Red Flag Amount": flagged_amount,
                "Status": "REVIEW REQUIRED" if n_flags > 0 else "CLEAN",
            })
            for _, r in df[df["Red_Flag"]].iterrows():
                all_flags.append({"File": entry["rel_path"], "LAN (folder name)": lan, "Borrower": borrower_name,
                                   "Date": r["sDate"].date(), "Amount": r["sAmount"], "Narration": r["sNarration"],
                                   "Counterparty": r.get("counter_party", ""), "Flag Reason": r["Flag_Reason"]})
            for bucket, amt in res["post"].groupby("End_Use_Bucket")["sAmount"].sum().items():
                bucket_totals[bucket] = bucket_totals.get(bucket, 0) + amt

        except Exception as e:
            errors.append({"File": entry["rel_path"], "Error": str(e)})

        if i % 25 == 0 or i == len(files):
            print(f"  processed {i}/{len(files)}  (errors so far: {len(errors)})")

    out_path = os.path.join(base_folder, f"{folder_label}_EUF_Report.xlsx")
    write_report(results, all_flags, bucket_totals, errors, out_path, len(files))

    print(f"\nDone. Report saved to: {out_path}")
    print(f"  Loans processed OK : {len(results)}")
    print(f"  Files with errors  : {len(errors)}")
    print(f"  Loans flagged      : {sum(1 for r in results if r['Status'] == 'REVIEW REQUIRED')}")


if __name__ == "__main__":
    main()
