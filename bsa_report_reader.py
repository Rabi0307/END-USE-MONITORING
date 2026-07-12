"""
bsa_report_reader.py
---------------------
Parses a full, multi-sheet Bank Statement Analyzer (BSA) report -- the kind
with sheets like 'Customer & Account Information', 'Overall Analysis',
'Fixed Obligations', 'Emi Transactions', 'AllAccountXns', etc. (the format
of report3_...xlsx) -- into a single structured summary you can use to
auto-fill the EUF Loan Master fields and add existing-debt / FOIR context
that a plain transaction list doesn't have.

This is a READER only -- it doesn't do end-use classification itself.
Pair it with end_use_engine.py (which already knows how to pull the
transaction-level sheet straight out of the same report) for the full
EUF analysis. See build_from_bsa_report.py for the combined pipeline.
"""

import pandas as pd
import openpyxl


def _sheet_to_dict(wb, sheet_name):
    """For 2-column 'Metric/Amount' or 'Items/Details' style sheets."""
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    return {r[0]: r[1] for r in rows[1:] if r and r[0] is not None}


def _sheet_to_df(path, sheet_name, wb=None):
    if wb is not None and sheet_name not in wb.sheetnames:
        return pd.DataFrame()
    try:
        return pd.read_excel(path, sheet_name=sheet_name)
    except ValueError:
        return pd.DataFrame()


def read_bsa_report(path: str) -> dict:
    """
    Returns a dict with:
      customer            -- dict from 'Customer & Account Information'
      overall_analysis     -- dict from 'Overall Analysis'
                              (Salary Credits, Fixed Obligations, EMI Payments,
                               FOIR, Other Inflows/Outflows, etc.)
      monthwise            -- DataFrame from 'Monthwise Credits and Debits'
      fixed_obligations     -- DataFrame from 'Fixed Obligations' (existing
                              recurring committed payments -- other EMIs,
                              rent, subscriptions, etc. detected in the
                              statement, independent of the loan being
                              tracked)
      emi_transactions      -- DataFrame from 'Emi Transactions'
      top_funds_transferred_by_category -- DataFrame, monthly category-wise
                              outflows as already computed by the BSA tool
      top_funds_received_by_category    -- DataFrame, same for inflows
      recurring_debits / recurring_credits -- DataFrames of repeating payments
    """
    wb = openpyxl.load_workbook(path, data_only=True)

    out = {
        "customer": _sheet_to_dict(wb, "Customer & Account Information"),
        "overall_analysis": _sheet_to_dict(wb, "Overall Analysis"),
        "monthwise": _sheet_to_df(path, "Monthwise Credits and Debits", wb),
        "fixed_obligations": _sheet_to_df(path, "Fixed Obligations", wb),
        "emi_transactions": _sheet_to_df(path, "Emi Transactions", wb),
        "top_funds_transferred_by_category": _sheet_to_df(path, "TopFundsTransferred by Category", wb),
        "top_funds_received_by_category": _sheet_to_df(path, "TopFundsReceived by Category", wb),
        "recurring_debits": _sheet_to_df(path, "Recurring Debits", wb),
        "recurring_credits": _sheet_to_df(path, "Recurring Credits", wb),
    }
    return out


def summarize_existing_obligations(bsa: dict) -> dict:
    """
    Pulls out the numbers most relevant to end-use / credit risk review:
    is there already EMI debt running on this account, how much, FOIR, etc.
    Useful red-flag context: if a large chunk of a NEW disbursement's outflow
    lands in 'Debt Servicing', cross-check here whether that's pre-existing
    (already running before disbursement) or new obligations that appeared
    only after disbursement (the latter is a stronger diversion signal).
    """
    oa = bsa.get("overall_analysis", {})
    cust = bsa.get("customer", {})
    fixed_ob = bsa.get("fixed_obligations", pd.DataFrame())

    existing_emi_total = 0.0
    if not fixed_ob.empty and "predictCategory" in fixed_ob.columns:
        emi_rows = fixed_ob[fixed_ob["predictCategory"].astype(str).str.contains("Emi", case=False, na=False)]
        amt_col = "amount" if "amount" in fixed_ob.columns else "sAmount"
        if amt_col in emi_rows.columns:
            existing_emi_total = float(emi_rows[amt_col].sum())

    return {
        "borrower_name": cust.get("sName", ""),
        "account_no": cust.get("sAccountNo", ""),
        "bank": cust.get("sBank", ""),
        "account_type": cust.get("sAccountType", ""),
        "current_balance": cust.get("sCurrentBalance", None),
        "monthly_avg_balance": cust.get("Monthly_Average_Balance", None),
        "opening_balance": cust.get("Opening Balance", None),
        "analysis_period": f"{cust.get('sAnalysisStartDate','')} to {cust.get('sAnalysisEndDate','')}",
        "emi_detected": cust.get("EMI Detected", None),
        "cheque_bounce": cust.get("Cheque Bounce", None),
        "foir": oa.get("FOIR", None),
        "fixed_obligations_total": oa.get("Fixed Obligations", None),
        "emi_payments_total": oa.get("EMI Payments", None),
        "salary_credits": oa.get("Salary Credits", None),
        "non_salary_income": oa.get("Non Salary Income", None),
        "existing_emi_from_fixed_obligations_sheet": existing_emi_total,
    }


if __name__ == "__main__":
    import argparse
    import json

    p = argparse.ArgumentParser(description="Read a full BSA report and print its key metrics")
    p.add_argument("--input", required=True)
    args = p.parse_args()

    bsa = read_bsa_report(args.input)
    summary = summarize_existing_obligations(bsa)
    print(json.dumps(summary, indent=2, default=str))
    print("\nFixed Obligations sample:")
    print(bsa["fixed_obligations"].head(10).to_string() if not bsa["fixed_obligations"].empty else "(none)")
